#!/usr/bin/env python3
"""
Import sales orders from a local Google Sheets export (.xlsx) into mini-erp DB.

HOW TO GET THE FILE
-------------------
1. Open your Google Sheet in the browser
2. File → Download → Microsoft Excel (.xlsx)
3. Save the file anywhere you like (default: scripts/sales_data.xlsx)

HOW TO RUN
----------
From mini-erp-backend/:

    uv sync
    uv run python scripts/import_sales_from_sheets.py

    # custom file location
    uv run python scripts/import_sales_from_sheets.py --file /path/to/export.xlsx

    # preview without writing to DB
    uv run python scripts/import_sales_from_sheets.py --dry-run

    # process only one month
    uv run python scripts/import_sales_from_sheets.py --tab "Sales November 2025"

WHAT IT DOES
------------
- Reads all monthly tabs (Sales May 2025 → Sales May 2026)
- Groups rows by Nomor Paket (package number) → one SalesOrder per package
- Maps Channel - Nama Toko to source_platform (SHOPEE or TIKTOK)
- Inserts SalesOrder + SalesOrderItem; safe to re-run (skips duplicates)
- COGS fields are left at 0 — fill them in a separate pass once purchase
  orders and FIFO layers are imported
"""

import argparse
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed.\nRun: uv sync   (or: pip install openpyxl)")

import psycopg

# ── Configuration ──────────────────────────────────────────────────────────────

DB = dict(dbname="mini_erp", user="postgres", password="postgres",
          host="localhost", port=5433)

DEFAULT_XLSX = Path("scripts/sales_data.xlsx")

WIB = ZoneInfo("Asia/Jakarta")
NS  = uuid.NAMESPACE_OID

MONTHLY_TABS = [
    "Sales May 2025",
    "Sales June 2025",
    "Sales July 2025",
    "Sales August 2025",
    "Sales September 2025",
    "Sales October 2025",
    "Sales November 2025",
    "Sales December 2025",
    "Sales January 2026",
    "Sales February 2026",
    "Sales March 2026",
    "Sales April 2026",
    "Sales May 2026",
]

# channel name (lowercase) → (source_platform, marketplace display name)
# Sora Kids is the old store name — treated as the same Shopee - Mirako Kids marketplace.
# All TikTok variants collapse to a single TikTok - MirakoKids marketplace.
CHANNEL_MAP: dict[str, tuple[str, str]] = {
    "shopee - mirako kids":         ("SHOPEE", "Shopee - Mirako Kids"),
    "shopee - sora kids":           ("SHOPEE", "Shopee - Mirako Kids"),
    "tiktok - mirakokids":          ("TIKTOK", "TikTok - MirakoKids"),
    "tiktok shop - mirakokids":     ("TIKTOK", "TikTok - MirakoKids"),
    "tiktok shop":                  ("TIKTOK", "TikTok - MirakoKids"),
}

# Desty order status → ERP OrderStatus
STATUS_MAP: dict[str, str] = {
    "completed":     "COMPLETED",
    "delivered":     "DELIVERED",
    "in_delivery":   "SHIPPING",
    "cancellations": "CANCELLED",
    "cancelled":     "CANCELLED",
    "pending":       "PENDING",
    "confirmed":     "CONFIRMED",
    "process":       "CONFIRMED",
    "returns":       "RETURNED",
    "returned":      "RETURNED",
    "refunded":      "RETURNED",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def uid(seed: str) -> str:
    """Deterministic uuid5 — same seed always → same ID; safe to re-run."""
    return str(uuid.uuid5(NS, f"mirako-kids:{seed}"))


def to_str(val: object) -> str:
    """Convert any openpyxl cell value to a stripped string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val).strip()


def parse_idr(raw: str | None) -> int:
    """Parse an IDR amount string → integer (handles commas, dots, spaces)."""
    if not raw:
        return 0
    clean = re.sub(r"[^\d]", "", str(raw).strip())
    return int(clean) if clean else 0


def parse_dt(raw: str | None) -> datetime | None:
    """Parse a date/time string → WIB-aware datetime."""
    if not raw or not str(raw).strip():
        return None
    s = str(raw).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=WIB)
        except ValueError:
            continue
    return None


def parse_province_city(address: str) -> tuple[str, str]:
    """
    Extract province and city from the Desty full address string.

    Desty format tail: ..., CITY, DISTRICT, PROVINCE, ID, POSTAL
    Example: "Jl. Abc, KOTA JAKARTA PUSAT, TANAH ABANG, DKI JAKARTA, ID, 10250"
    Returns : ("Dki Jakarta", "Kota Jakarta Pusat")
    """
    if not address:
        return "", ""
    parts = [p.strip() for p in address.split(",")]
    parts = [p for p in parts if p and p != "ID" and not re.match(r"^\d{4,6}$", p)]
    if len(parts) >= 3:
        return parts[-1].title(), parts[-3].title()
    if len(parts) == 2:
        return parts[-1].title(), parts[-2].title()
    return "", ""


# ── XLSX reading ───────────────────────────────────────────────────────────────

def read_xlsx_tab(wb: openpyxl.Workbook, tab_name: str) -> list[dict]:
    """
    Return all non-blank rows from a worksheet as dicts keyed by header name.
    openpyxl may return Python datetime objects for date cells — to_str() handles them.
    """
    if tab_name not in wb.sheetnames:
        return []

    ws = wb[tab_name]
    rows_iter = ws.iter_rows(values_only=True)

    # First non-empty row = headers
    headers: list[str] = []
    for raw in rows_iter:
        headers = [to_str(c) for c in raw]
        if any(headers):
            break

    if not headers:
        return []

    result = []
    for raw in rows_iter:
        if not any(c is not None and to_str(c) for c in raw):
            continue
        padded = list(raw) + [None] * max(0, len(headers) - len(raw))
        result.append({headers[i]: to_str(padded[i]) for i in range(len(headers))})

    return result


# ── DB helpers ─────────────────────────────────────────────────────────────────

def load_refs(cur: psycopg.Cursor) -> dict:
    """Load company, warehouse, marketplace, and variant IDs from DB."""
    cur.execute("SELECT company_id::text FROM core_company LIMIT 1")
    row = cur.fetchone()
    if not row:
        raise RuntimeError("No company in DB — run import_master_data.py first")
    company_id = row[0]

    cur.execute("""
        SELECT warehouse_id::text FROM inventory_warehouse
        WHERE is_active = TRUE ORDER BY cdate LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        raise RuntimeError("No active warehouse in DB — run import_master_data.py first")
    warehouse_id = row[0]

    cur.execute("SELECT marketplace_id::text, name FROM core_marketplace")
    marketplaces: dict[str, str] = {r[1].lower(): r[0] for r in cur.fetchall()}

    cur.execute("SELECT product_variant_id::text, sku_variant_code FROM inventory_productvariant")
    variants: dict[str, str] = {r[1]: r[0] for r in cur.fetchall()}

    return {
        "company_id":  company_id,
        "warehouse_id": warehouse_id,
        "marketplaces": marketplaces,
        "variants":    variants,
        "unknown_skus": set(),
    }


def ensure_marketplace(
    cur: psycopg.Cursor,
    channel_raw: str,
    refs: dict,
    now: datetime,
    dry_run: bool,
) -> str | None:
    """Return marketplace_id for a channel string, creating the row when needed."""
    info = CHANNEL_MAP.get(channel_raw.strip().lower())
    if not info:
        if channel_raw:
            print(f"  WARNING: unknown channel {channel_raw!r} — no marketplace assigned")
        return None
    _, mp_name = info
    key = mp_name.lower()
    if key in refs["marketplaces"]:
        return refs["marketplaces"][key]
    mp_id = uid(f"marketplace:{key}")
    if not dry_run:
        cur.execute("""
            INSERT INTO core_marketplace
                (marketplace_id, name, url, status, connected_time,
                 is_active, shipping_config, cdate, udate)
            VALUES (%s::uuid, %s, NULL, 'active', NULL, TRUE, '{}'::jsonb, %s, %s)
            ON CONFLICT (marketplace_id) DO NOTHING
        """, (mp_id, mp_name, now, now))
    refs["marketplaces"][key] = mp_id
    print(f"  [+] Marketplace created: {mp_name}")
    return mp_id


# ── Order grouping ─────────────────────────────────────────────────────────────

def group_rows_into_orders(rows: list[dict]) -> dict[str, dict]:
    """
    Group Sheet rows by Nomor Paket → one SalesOrder per package.
    Falls back to Nomor Pesanan (di Marketplace) for cancelled/no-tracking orders.
    Items with the same SKU in the same order are merged (qty summed).
    """
    orders: dict[str, dict] = {}

    for row in rows:
        pkg_num = row.get("Nomor Paket", "").strip()
        mkt_num = row.get("Nomor Pesanan (di Marketplace)", "").strip()
        order_key = pkg_num or mkt_num
        if not order_key:
            continue

        if order_key not in orders:
            channel     = row.get("Channel - Nama Toko", "").strip()
            platform    = CHANNEL_MAP.get(channel.lower(), ("SHOPEE", ""))[0]
            status_raw  = row.get("Status Pesanan", "").strip().lower()
            status      = STATUS_MAP.get(status_raw, "COMPLETED")
            address     = row.get("Alamat Pembeli", "").strip()
            province, city = parse_province_city(address)

            orders[order_key] = {
                "package_number":           pkg_num,
                "marketplace_order_id":     mkt_num or pkg_num,
                "marketplace_order_number": pkg_num or mkt_num,
                "channel":                  channel,
                "source_platform":          platform,
                "status":                   status,
                "order_date":               parse_dt(row.get("Tanggal Pesanan Dibuat")),
                "customer_name":            row.get("Nama Pembeli", "").strip(),
                "customer_phone":           row.get("Nomor Telepon Pembeli", "").strip(),
                "shipping_address":         address,
                "shipping_province":        province,
                "shipping_city":            city,
                "courier_name":             row.get("Kurir", "").strip(),
                "tracking_number":          row.get("Nomor AWB/Resi", "").strip(),
                "shipping_fee":             parse_idr(row.get("Biaya Pengiriman Final")),
                "items": {},  # sku → item dict; merged by SKU within same order
            }

        sku = row.get("SKU Master", "").strip()
        if not sku:
            continue

        unit_price  = parse_idr(row.get("Harga Satuan"))
        paid_price  = parse_idr(row.get("Harga Dibayar")) or unit_price
        item_qty    = max(parse_idr(row.get("Jumlah")), 1)
        service_fee = parse_idr(row.get("Biaya Layanan"))
        discount    = max(0, unit_price - paid_price) * item_qty
        line_total  = parse_idr(row.get("Subtotal Produk")) or (paid_price * item_qty)

        if sku in orders[order_key]["items"]:
            it = orders[order_key]["items"][sku]
            it["quantity"]               += item_qty
            it["discount_amount"]        += discount
            it["service_fee"]            += service_fee
            it["total_marketplace_fee"]  += service_fee
            it["line_total"]             += line_total
        else:
            orders[order_key]["items"][sku] = {
                "sku":                  sku,
                "quantity":             item_qty,
                "selling_price":        unit_price,
                "discount_amount":      discount,
                "service_fee":          service_fee,
                "total_marketplace_fee": service_fee,
                "line_total":           line_total,
            }

    return orders


# ── DB write ───────────────────────────────────────────────────────────────────

def insert_orders(
    cur: psycopg.Cursor,
    orders: dict[str, dict],
    refs: dict,
    now: datetime,
    dry_run: bool,
) -> tuple[int, int, int]:
    """
    Upsert SalesOrder + SalesOrderItem rows.
    Returns (inserted, skipped_already_exists, skipped_no_valid_items).
    """
    inserted = skipped_existing = skipped_no_items = 0

    for order_key, order in orders.items():
        label        = order["package_number"] or order["marketplace_order_id"]
        order_number = f"SO-{label}"

        cur.execute("SELECT 1 FROM sales_salesorder WHERE order_number = %s", (order_number,))
        if cur.fetchone():
            skipped_existing += 1
            continue

        valid_items: list[dict] = []
        for it in order["items"].values():
            variant_id = refs["variants"].get(it["sku"])
            if not variant_id:
                refs["unknown_skus"].add(it["sku"])
                continue
            valid_items.append({**it, "product_variant_id": variant_id})

        if not valid_items and order["status"] not in ("CANCELLED", "RETURNED"):
            skipped_no_items += 1
            continue

        subtotal              = sum(i["line_total"]      for i in valid_items)
        total_discount        = sum(i["discount_amount"] for i in valid_items)
        total_marketplace_fee = sum(i["service_fee"]     for i in valid_items)
        net_revenue           = subtotal - total_discount - total_marketplace_fee

        marketplace_id = ensure_marketplace(cur, order["channel"], refs, now, dry_run)
        so_id          = uid(f"so:{order_key}")
        order_date     = order["order_date"] or now

        if not dry_run:
            cur.execute("""
                INSERT INTO sales_salesorder (
                    sales_order_id, company_id, order_number,
                    marketplace_id, marketplace_order_id, marketplace_order_number,
                    status, source_platform, warehouse_id,
                    customer_name, customer_phone, shipping_address,
                    shipping_province, shipping_city,
                    courier_name, tracking_number,
                    shipping_fee, shipping_fee_seller,
                    subtotal, total_discount, total_marketplace_fee,
                    total_cogs, net_revenue, gross_profit,
                    order_date, note, cdate, udate
                ) VALUES (
                    %s::uuid, %s::uuid, %s,
                    %s::uuid, %s, %s,
                    %s, %s, %s::uuid,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s
                ) ON CONFLICT (order_number) DO NOTHING
            """, (
                so_id, refs["company_id"], order_number,
                marketplace_id, order["marketplace_order_id"], order["marketplace_order_number"],
                order["status"], order["source_platform"], refs["warehouse_id"],
                order["customer_name"], order["customer_phone"], order["shipping_address"],
                order["shipping_province"], order["shipping_city"],
                order["courier_name"], order["tracking_number"],
                order["shipping_fee"], 0,
                subtotal, total_discount, total_marketplace_fee,
                0, net_revenue, 0,
                order_date, "", now, now,
            ))

            for item in valid_items:
                item_id = uid(f"so-item:{order_key}:{item['sku']}")
                cur.execute("""
                    INSERT INTO sales_salesorderitem (
                        sales_order_item_id, company_id, sales_order_id,
                        product_variant_id, quantity,
                        selling_price, discount_amount, commission_fee,
                        service_fee, total_marketplace_fee,
                        actual_cogs_per_unit, actual_cogs_total,
                        line_total, cdate, udate
                    ) VALUES (
                        %s::uuid, %s::uuid, %s::uuid,
                        %s::uuid, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s, %s
                    ) ON CONFLICT (sales_order_item_id) DO NOTHING
                """, (
                    item_id, refs["company_id"], so_id,
                    item["product_variant_id"], item["quantity"],
                    item["selling_price"], item["discount_amount"], 0,
                    item["service_fee"], item["total_marketplace_fee"],
                    0, 0,
                    item["line_total"], now, now,
                ))

        inserted += 1

    return inserted, skipped_existing, skipped_no_items


# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import Desty/Google Sheets sales orders into mini-erp DB"
    )
    parser.add_argument(
        "--file", type=Path, default=DEFAULT_XLSX,
        metavar="PATH",
        help=f"Path to the .xlsx export from Google Sheets (default: {DEFAULT_XLSX})",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Show counts without writing anything to the DB",
    )
    parser.add_argument(
        "--tab", metavar="TAB_NAME",
        help="Process only one tab, e.g. 'Sales November 2025'",
    )
    args = parser.parse_args()

    xlsx_path: Path = args.file
    if not xlsx_path.exists():
        sys.exit(
            f"File not found: {xlsx_path}\n\n"
            "To get the file:\n"
            "  1. Open your Google Sheet in Chrome\n"
            "  2. File → Download → Microsoft Excel (.xlsx)\n"
            f"  3. Save it as {DEFAULT_XLSX}  (or pass --file <path>)\n"
        )

    tabs = [args.tab] if args.tab else MONTHLY_TABS

    print(f"Loading workbook: {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}\n")

    now = datetime.now(timezone.utc)

    print("Connecting to database ...")
    conn = psycopg.connect(**DB)
    cur  = conn.cursor()

    try:
        refs = load_refs(cur)
        print(f"  Company:   {refs['company_id']}")
        print(f"  Warehouse: {refs['warehouse_id']}")
        print(f"  Variants:  {len(refs['variants'])} loaded")
        if args.dry_run:
            print("  [DRY RUN — nothing will be written]\n")

        total_inserted = total_existing = total_no_items = 0

        for tab_name in tabs:
            print(f"\n── {tab_name} {'─' * max(0, 45 - len(tab_name))}")
            rows = read_xlsx_tab(wb, tab_name)
            if not rows:
                print("  (tab not found or empty — skipping)")
                continue
            print(f"  Sheet rows: {len(rows)}")

            orders = group_rows_into_orders(rows)
            print(f"  Orders:     {len(orders)}")

            ins, ex, ni = insert_orders(cur, orders, refs, now, args.dry_run)
            total_inserted  += ins
            total_existing  += ex
            total_no_items  += ni
            print(f"  Inserted: {ins}  |  Already exists: {ex}  |  No valid items: {ni}")

        if not args.dry_run:
            conn.commit()

    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()
        wb.close()

    mode = "DRY RUN" if args.dry_run else "DONE"
    print(f"\n{'─' * 55}")
    print(f"{mode}")
    print(f"  Orders inserted:              {total_inserted}")
    print(f"  Orders already in DB:         {total_existing}")
    print(f"  Orders skipped (no SKU match): {total_no_items}")

    if refs["unknown_skus"]:
        print(f"\n  Unrecognised SKUs ({len(refs['unknown_skus'])}) — not in product_variant table:")
        for sku in sorted(refs["unknown_skus"]):
            print(f"    {sku}")
        print("\n  Add missing variants via import_master_data.py, then re-run this script.")


if __name__ == "__main__":
    main()
