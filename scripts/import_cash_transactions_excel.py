"""
Phase 5 — Import Finance sheet from Mirakokids FinOps.xlsx into CashTransaction records.

Run from mini-erp-backend/:
    uv run python scripts/import_cash_transactions_excel.py
"""

from datetime import datetime, timezone

import openpyxl
import psycopg
import ulid as ulid_lib

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = "/Users/jtf01644/personal/mini-erp-project/data/Mirakokids FinOps.xlsx"
SHEET_NAME = "Finance"

DB_PARAMS = dict(
    host="localhost",
    port=5433,
    dbname="mini_erp",
    user="postgres",
    password="postgres",
)

CATEGORY_MAP: dict[str, str] = {
    "In (In)": "EQUITY_INJECTION",
    "In (Sales)": "SALES_SETTLEMENT",
    "In (Others)": "OTHER_INCOME",
    "Out (Order)": "OTHER_EXPENSE",
    "Out (Ops)": "OTHER_EXPENSE",
    "Out (Marketing)": "OTHER_EXPENSE",
    "Out (Shipping)": "OTHER_EXPENSE",
    "Out (Salary)": "OTHER_EXPENSE",
    "Out (Others)": "OTHER_EXPENSE",
}


def new_ulid() -> str:
    return str(ulid_lib.new().uuid)


def now() -> datetime:
    return datetime.now(timezone.utc)


def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{SHEET_NAME}' not found in workbook")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    print(f"Sheet '{SHEET_NAME}' has {len(rows)} rows (including header)")

    print("\nConnecting to database...")
    with psycopg.connect(**DB_PARAMS, autocommit=True) as conn:
        with conn.cursor() as cur:
            # ------------------------------------------------------------------
            # STEP 0 — Bootstrap
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 0: Bootstrap")
            print("=" * 60)

            cur.execute("SELECT company_id FROM core_company WHERE name = 'Mirako' LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise RuntimeError("Company 'Mirako' not found")
            company_id = str(row[0])
            print(f"  Company 'Mirako' id: {company_id}")

            # ------------------------------------------------------------------
            # STEP 1 — Idempotency Reset
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 1: Idempotency Reset")
            print("=" * 60)

            cur.execute(
                "DELETE FROM finance_cashtransaction WHERE company_id = %s",
                (company_id,),
            )
            print(f"  Deleted existing CashTransaction records: {cur.rowcount}")

            # ------------------------------------------------------------------
            # STEP 2 — Read and import rows
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 2: Import Rows")
            print("=" * 60)

            inserted = 0
            skipped_no_num = 0
            skipped_no_date = 0
            skipped_no_value = 0

            for row_idx, data_row in enumerate(rows[1:], start=2):
                # col 1: row number
                row_num = data_row[0] if len(data_row) > 0 else None
                if row_num is None:
                    skipped_no_num += 1
                    continue
                try:
                    int(row_num)
                except (ValueError, TypeError):
                    skipped_no_num += 1
                    continue

                # col 3: Date
                raw_date = data_row[2] if len(data_row) > 2 else None
                if raw_date is None:
                    skipped_no_date += 1
                    continue
                if isinstance(raw_date, datetime):
                    transaction_date = raw_date.date()
                else:
                    transaction_date = raw_date

                # col 4: Description
                raw_desc = data_row[3] if len(data_row) > 3 else None
                description = str(raw_desc).strip()[:2000] if raw_desc else ""

                # col 5: Type1 ("In" or "Out")
                raw_type1 = data_row[4] if len(data_row) > 4 else None
                type1 = str(raw_type1).strip() if raw_type1 else "Out"

                # col 6: Type (category string)
                raw_type = data_row[5] if len(data_row) > 5 else None
                type_str = str(raw_type).strip() if raw_type else ""

                # col 7: Value (IDR)
                raw_value = data_row[6] if len(data_row) > 6 else None
                if raw_value is None:
                    skipped_no_value += 1
                    continue
                try:
                    value_float = float(raw_value)
                except (ValueError, TypeError):
                    skipped_no_value += 1
                    continue
                amount = int(abs(value_float))

                # Derive fields
                transaction_type = "INFLOW" if type1 == "In" else "OUTFLOW"
                category = CATEGORY_MAP.get(type_str, "OTHER_EXPENSE")
                note = f"Excel Type: {type_str}"

                cash_transaction_id = new_ulid()
                ts = now()

                cur.execute(
                    """
                    INSERT INTO finance_cashtransaction (
                        cash_transaction_id, company_id,
                        transaction_date, description, amount,
                        transaction_type, category,
                        reference_number, note,
                        cdate, udate
                    ) VALUES (
                        %s, %s,
                        %s, %s, %s,
                        %s, %s,
                        %s, %s,
                        %s, %s
                    )
                    """,
                    (
                        cash_transaction_id,
                        company_id,
                        transaction_date,
                        description,
                        amount,
                        transaction_type,
                        category,
                        "",
                        note,
                        ts,
                        ts,
                    ),
                )
                inserted += 1

            print(f"  Inserted: {inserted}")
            print(f"  Skipped (no row number): {skipped_no_num}")
            print(f"  Skipped (no date): {skipped_no_date}")
            print(f"  Skipped (no/zero value): {skipped_no_value}")

            # ------------------------------------------------------------------
            # STEP 3 — Summary
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 3: Summary")
            print("=" * 60)

            cur.execute(
                """
                SELECT COUNT(*),
                       COALESCE(SUM(CASE WHEN transaction_type = 'INFLOW' THEN amount ELSE 0 END), 0),
                       COALESCE(SUM(CASE WHEN transaction_type = 'OUTFLOW' THEN amount ELSE 0 END), 0)
                FROM finance_cashtransaction WHERE company_id = %s
                """,
                (company_id,),
            )
            total_count, total_inflow, total_outflow = cur.fetchone()
            total_inflow = int(total_inflow)
            total_outflow = int(total_outflow)

            print(f"  Total records: {total_count}")
            print(f"  Total INFLOW:  {total_inflow} IDR")
            print(f"  Total OUTFLOW: {total_outflow} IDR")
            print(f"  Net balance:   {total_inflow - total_outflow} IDR")

            # ------------------------------------------------------------------
            # Validations
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("VALIDATIONS")
            print("=" * 60)

            assert total_count > 0, "No records imported!"
            print(f"  [PASS] {total_count} records > 0")

            assert total_inflow > total_outflow, (
                f"Total inflow ({total_inflow}) should exceed total outflow ({total_outflow})"
            )
            print(f"  [PASS] INFLOW ({total_inflow}) > OUTFLOW ({total_outflow})")

            cur.execute(
                "SELECT COUNT(*) FROM finance_cashtransaction WHERE company_id = %s AND amount = 0",
                (company_id,),
            )
            zero_count = cur.fetchone()[0]
            assert zero_count == 0, f"Found {zero_count} records with amount = 0"
            print("  [PASS] No zero-amount records")

            # Spot-check first row
            cur.execute(
                """
                SELECT transaction_date, description, amount, transaction_type, category
                FROM finance_cashtransaction
                WHERE company_id = %s
                ORDER BY transaction_date ASC, cdate ASC
                LIMIT 1
                """,
                (company_id,),
            )
            first = cur.fetchone()
            assert first is not None, "No records to spot-check"
            td, desc, amt, ttype, cat = first
            expected_date = td  # just compare as date object
            assert str(expected_date) == "2025-04-08", (
                f"First row date expected 2025-04-08, got {expected_date}"
            )
            assert desc == "Chip In Albert", (
                f"First row desc expected 'Chip In Albert', got '{desc}'"
            )
            assert amt == 25000000, f"First row amount expected 25000000, got {amt}"
            assert ttype == "INFLOW", f"First row type expected INFLOW, got {ttype}"
            assert cat == "EQUITY_INJECTION", (
                f"First row category expected EQUITY_INJECTION, got {cat}"
            )
            print(
                f"  [PASS] Spot-check: first row = {expected_date}, {desc}, {amt}, {ttype}, {cat}"
            )

            print("\n" + "=" * 60)
            print("ALL VALIDATIONS PASSED")
            print("=" * 60)


if __name__ == "__main__":
    main()
