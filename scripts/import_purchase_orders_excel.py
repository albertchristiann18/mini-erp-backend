"""
Phase 3 — Import all 14 historical purchase orders with FIFO batches,
stock movements, and warehouse stock from Excel.

Run from mini-erp-backend/:
    uv run python scripts/import_purchase_orders_excel.py
"""

import re
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal

import openpyxl
import psycopg
import ulid as ulid_lib

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = (
    "/Users/jtf01644/personal/mini-erp-project/data/Purchase Order Detail Overview (1).xlsx"
)

DB_PARAMS = dict(
    host="localhost",
    port=5433,
    dbname="mini_erp",
    user="postgres",
    password="postgres",
)

SHEETS_TO_IMPORT = [
    "PO Recap 2 April 2025",
    "PO Recap 2 Juni 2025",
    "PO Recap 19 Juni 2025",
    "Recap 11 Sept 2025",
    "Recap 13 Oct 2025",
    "Recap 5 Dec 2025",
    "Recap 21 Dec 2025",
    "Recap 9 Jan 2026",
    "Recap 24 Feb 2026",
    "Split Recap 24 Feb 2026",
    "Recap 29 Mar 2026",
    "Split Recap 29 Mar 2026",
    "Recap 12 June 2026",
    "Recap 29 June 2026",
]

MONTH_MAP = {
    "januari": 1,
    "january": 1,
    "jan": 1,
    "februari": 2,
    "february": 2,
    "feb": 2,
    "maret": 3,
    "march": 3,
    "mar": 3,
    "april": 4,
    "mei": 5,
    "may": 5,
    "juni": 6,
    "june": 6,
    "juli": 7,
    "july": 7,
    "agustus": 8,
    "august": 8,
    "september": 9,
    "sept": 9,
    "oktober": 10,
    "oct": 10,
    "october": 10,
    "november": 11,
    "nov": 11,
    "desember": 12,
    "december": 12,
    "dec": 12,
}

DATE_PATTERN = re.compile(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})")


def new_ulid() -> str:
    return str(ulid_lib.new().uuid)


def now() -> datetime:
    return datetime.now(timezone.utc)


def parse_po_date(sheet_name: str) -> date | None:
    m = DATE_PATTERN.search(sheet_name.lower())
    if not m:
        return None
    day = int(m.group(1))
    month_name = m.group(2)
    year = int(m.group(3))
    month = MONTH_MAP.get(month_name)
    if month is None:
        return None
    return date(year, month, day)


def detect_columns(headers_row: tuple) -> dict[str, int] | None:
    sku_col = None
    price_col = None
    disc_col = None
    order_col = None
    soh_col = None

    for idx, val in enumerate(headers_row):
        if val is None:
            continue
        key = str(val).strip().lower()

        if key == "sku code" and sku_col is None:
            sku_col = idx
        elif key == "price disc" and disc_col is None:
            disc_col = idx
        elif key == "order qty" and order_col is None:
            order_col = idx
        elif key == "soh" and soh_col is None:
            soh_col = idx
        elif (
            price_col is None
            and key.startswith("price")
            and "disc" not in key
            and "total" not in key
        ):
            price_col = idx

    if sku_col is None or order_col is None:
        return None

    return {
        "sku": sku_col,
        "price": price_col,
        "disc": disc_col,
        "order": order_col,
        "soh": soh_col,
    }


def extract_exchange_rate(row1: tuple) -> int | None:
    cells = list(row1)
    for i, val in enumerate(cells):
        if val is not None and str(val).strip().upper() == "RMB":
            for j in range(i + 1, len(cells)):
                if cells[j] is not None:
                    try:
                        return int(round(float(str(cells[j]))))
                    except (ValueError, TypeError):
                        pass
    return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s == "#REF!":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    f = safe_float(val)
    if f is None:
        return None
    return int(f)


def is_sku_valid(sku_code: str | None) -> bool:
    if sku_code is None:
        return False
    s = str(sku_code).strip()
    if s == "":
        return False
    if s.replace(".", "").replace("-", "").isdigit():
        return False
    if s.startswith("Total") or "#" in s:
        return False
    return True


def main():
    ts = now()

    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")

    print("\nConnecting to database...")
    with psycopg.connect(**DB_PARAMS, autocommit=True) as conn:
        with conn.cursor() as cur:
            # ------------------------------------------------------------------
            # Step 0 — Bootstrap: company, warehouse, supplier, variant_map
            # ------------------------------------------------------------------
            cur.execute("SELECT company_id FROM core_company WHERE name = 'Mirako' LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise RuntimeError("Company 'Mirako' not found")
            company_id = str(row[0])
            print(f"Company 'Mirako' id: {company_id}")

            cur.execute(
                "SELECT warehouse_id FROM inventory_warehouse WHERE company_id = %s ORDER BY cdate LIMIT 1",
                (company_id,),
            )
            row = cur.fetchone()
            if not row:
                raise RuntimeError("No warehouse found for company Mirako")
            warehouse_id = str(row[0])
            print(f"Warehouse id: {warehouse_id}")

            cur.execute(
                "SELECT supplier_id FROM inventory_supplier WHERE company_id = %s LIMIT 1",
                (company_id,),
            )
            row = cur.fetchone()
            if not row:
                raise RuntimeError("No supplier found for company Mirako")
            supplier_id = str(row[0])
            print(f"Supplier id: {supplier_id}")

            # Build variant_map: sku_variant_code → (product_variant_id, product_id)
            cur.execute(
                "SELECT product_variant_id, product_id, sku_variant_code "
                "FROM inventory_productvariant WHERE company_id = %s",
                (company_id,),
            )
            variant_map: dict[str, tuple[str, str]] = {}
            for pid, prod_id, sku in cur.fetchall():
                variant_map[str(sku)] = (str(pid), str(prod_id))
            print(f"Variant map built: {len(variant_map)} variants")

            # ------------------------------------------------------------------
            # Step 0b — Disable auto PO number trigger, delete existing HIST POs
            # ------------------------------------------------------------------
            print("\nDisabling trigger trg_generate_po_number ...")
            cur.execute(
                "ALTER TABLE purchasing_purchaseorder DISABLE TRIGGER trg_generate_po_number"
            )

            print("Deleting existing HIST purchase orders (idempotency)...")
            cur.execute(
                "DELETE FROM purchasing_purchaseorderdetail "
                "WHERE purchase_order_id IN ("
                "  SELECT purchase_order_id FROM purchasing_purchaseorder "
                "  WHERE company_id = %s AND purchase_order_number LIKE 'HIST-%%')",
                (company_id,),
            )
            deleted_details = cur.rowcount
            cur.execute(
                "DELETE FROM purchasing_purchaseorder "
                "WHERE company_id = %s AND purchase_order_number LIKE 'HIST-%%'",
                (company_id,),
            )
            deleted_pos = cur.rowcount
            print(f"  Deleted {deleted_details} detail rows, {deleted_pos} PO headers")

            # ------------------------------------------------------------------
            # Step 1 — Reset stock
            # ------------------------------------------------------------------
            print("\nResetting stock for Mirako variants...")
            cur.execute(
                "UPDATE inventory_productvariantwarehouse "
                "SET incoming_qty=0, outgoing_qty=0, physical_qty=0, checkout_qty=0 "
                "WHERE product_variant_id IN ("
                "  SELECT product_variant_id FROM inventory_productvariant WHERE company_id=%s)",
                (company_id,),
            )
            cur.execute(
                "UPDATE inventory_productvariant "
                "SET total_incoming_qty=0, total_outgoing_qty=0, total_available_qty=0 "
                "WHERE company_id=%s",
                (company_id,),
            )
            cur.execute(
                "DELETE FROM inventory_stockmovement "
                "WHERE product_variant_id IN ("
                "  SELECT product_variant_id FROM inventory_productvariant WHERE company_id=%s)",
                (company_id,),
            )
            cur.execute(
                "DELETE FROM inventory_productcogs "
                "WHERE product_variant_id IN ("
                "  SELECT product_variant_id FROM inventory_productvariant WHERE company_id=%s)",
                (company_id,),
            )
            print("Stock reset complete.")

            # ------------------------------------------------------------------
            # Step 2 — Import each PO sheet
            # ------------------------------------------------------------------
            po_year_counter: dict[int, int] = defaultdict(int)
            total_skipped_sheets = 0
            # Must persist across sheets so balance chain is continuous for SKUs ordered multiple times
            current_physical_qty: dict[str, int] = defaultdict(int)

            print("\n" + "=" * 60)
            print("IMPORTING PURCHASE ORDERS")
            print("=" * 60)

            for sheet_name in SHEETS_TO_IMPORT:
                if sheet_name not in wb.sheetnames:
                    print(f"  WARNING: Sheet '{sheet_name}' not found — skipping")
                    total_skipped_sheets += 1
                    continue

                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    print(f"  WARNING: Sheet '{sheet_name}' has too few rows — skipping")
                    total_skipped_sheets += 1
                    continue

                row1 = rows[0]
                row2 = rows[1]

                # Check row 1 contains "RMB" — skip sheets that don't
                has_rmb = False
                for val in row1:
                    if val is not None and "rmb" in str(val).strip().lower():
                        has_rmb = True
                        break
                if not has_rmb:
                    print(
                        f"  WARNING: Sheet '{sheet_name}' row 1 does not contain 'RMB' — skipping"
                    )
                    total_skipped_sheets += 1
                    continue

                # Detect columns
                cols = detect_columns(row2)
                if cols is None:
                    print(f"  WARNING: Sheet '{sheet_name}' missing required columns — skipping")
                    total_skipped_sheets += 1
                    continue

                sku_col = cols["sku"]
                price_col = cols["price"]
                disc_col = cols["disc"]
                order_col = cols["order"]
                soh_col = cols["soh"]

                # Exchange rate
                exchange_rate_int = extract_exchange_rate(row1)
                if exchange_rate_int is None:
                    print(
                        f"  WARNING: Sheet '{sheet_name}' could not determine exchange rate — skipping"
                    )
                    total_skipped_sheets += 1
                    continue

                # Parse PO date
                po_date = parse_po_date(sheet_name)
                if po_date is None:
                    print(f"  WARNING: Sheet '{sheet_name}' could not parse date — skipping")
                    total_skipped_sheets += 1
                    continue

                year = po_date.year
                po_year_counter[year] += 1
                po_number = f"HIST-{year}-{po_year_counter[year]:03d}"

                # Create PurchaseOrder
                po_id = new_ulid()
                cur.execute(
                    """
                    INSERT INTO purchasing_purchaseorder (
                        purchase_order_id, company_id, purchase_order_number, status,
                        invoice_date, delivery_date,
                        warehouse_id, supplier_id, supplier_name,
                        currency, exchange_rate,
                        total_ordered_qty, total_received_qty,
                        total_item_amount, procure_amount,
                        shipping_fee, has_discount,
                        cdate, udate
                    ) VALUES (
                        %s, %s, %s, 'COMPLETED',
                        %s, %s,
                        %s, %s, 'Ancorelala',
                        'RMB', %s,
                        0, 0,
                        0, 0,
                        0, FALSE,
                        %s, %s
                    )
                    """,
                    (
                        po_id,
                        company_id,
                        po_number,
                        po_date,
                        po_date,
                        warehouse_id,
                        supplier_id,
                        exchange_rate_int,
                        ts,
                        ts,
                    ),
                )

                # Process data rows
                n_rows = 0
                n_skipped = 0
                po_total_qty = 0
                po_total_idr = 0

                last_valid_price_rmb: float | None = None

                for data_row in rows[2:]:
                    if len(data_row) <= max(sku_col, order_col):
                        n_skipped += 1
                        continue

                    raw_sku = data_row[sku_col]
                    if not is_sku_valid(raw_sku):
                        n_skipped += 1
                        continue
                    sku_code = str(raw_sku).strip()

                    # Order qty
                    order_qty_raw = data_row[order_col]
                    order_qty = safe_int(order_qty_raw)
                    if order_qty is None or order_qty <= 0:
                        n_skipped += 1
                        continue

                    # Price logic with carry-forward
                    disc = (
                        safe_float(data_row[disc_col])
                        if disc_col is not None and disc_col < len(data_row)
                        else None
                    )
                    price = (
                        safe_float(data_row[price_col])
                        if price_col is not None and price_col < len(data_row)
                        else None
                    )

                    effective_price_rmb: float | None = None
                    if disc is not None and disc > 1.0:
                        effective_price_rmb = disc
                    elif price is not None:
                        effective_price_rmb = price

                    if effective_price_rmb is not None:
                        last_valid_price_rmb = effective_price_rmb
                    else:
                        effective_price_rmb = last_valid_price_rmb

                    if effective_price_rmb is None or effective_price_rmb <= 0:
                        n_skipped += 1
                        continue

                    # Lookup variant
                    if sku_code not in variant_map:
                        print(f"  SKIP: {sku_code} not found in DB")
                        n_skipped += 1
                        continue
                    variant_id, _ = variant_map[sku_code]

                    # Calculate prices
                    unit_price_base = int(effective_price_rmb * exchange_rate_int)
                    total_price_foreign = effective_price_rmb * order_qty
                    total_price_base = int(effective_price_rmb * exchange_rate_int * order_qty)

                    # INSERT PurchaseOrderDetail
                    pod_id = new_ulid()
                    cur.execute(
                        """
                        INSERT INTO purchasing_purchaseorderdetail (
                            purchase_order_detail_id, company_id, purchase_order_id,
                            product_variant_id,
                            ordered_qty, received_qty, updated_qty, draft_product_name,
                            unit_price_foreign, unit_price_base,
                            total_price_foreign, total_price_base,
                            discounted_unit_price_foreign, discounted_unit_price_base,
                            discounted_total_price_foreign, discounted_total_price_base,
                            stock_on_hand, incoming_qty,
                            cdate, udate
                        ) VALUES (
                            %s, %s, %s,
                            %s,
                            %s, %s, 0, '',
                            %s, %s,
                            %s, %s,
                            %s, %s,
                            %s, %s,
                            %s, 0,
                            %s, %s
                        )
                        """,
                        (
                            pod_id,
                            company_id,
                            po_id,
                            variant_id,
                            order_qty,
                            order_qty,
                            Decimal(str(effective_price_rmb)),
                            unit_price_base,
                            Decimal(str(total_price_foreign)),
                            total_price_base,
                            Decimal(str(effective_price_rmb)),
                            unit_price_base,
                            Decimal(str(total_price_foreign)),
                            total_price_base,
                            safe_int(data_row[soh_col]) if soh_col is not None else 0,
                            ts,
                            ts,
                        ),
                    )

                    # INSERT ProductCogs (FIFO batch)
                    cogs_id = new_ulid()
                    cur.execute(
                        """
                        INSERT INTO inventory_productcogs (
                            product_cogs_id, company_id, product_variant_id, warehouse_id,
                            reference_number, purchase_date,
                            price_rmb, exchange_rate, cogs_amount,
                            allocated_shipping_fee, allocated_delivery_fee, allocated_commission_fee,
                            original_qty, remaining_qty, is_active,
                            cdate, udate
                        ) VALUES (
                            %s, %s, %s, %s,
                            %s, %s,
                            %s, %s, %s,
                            0, 0, 0,
                            %s, %s, TRUE,
                            %s, %s
                        )
                        """,
                        (
                            cogs_id,
                            company_id,
                            variant_id,
                            warehouse_id,
                            po_number,
                            po_date,
                            Decimal(str(effective_price_rmb)),
                            exchange_rate_int,
                            unit_price_base,
                            order_qty,
                            order_qty,
                            ts,
                            ts,
                        ),
                    )

                    # INSERT StockMovement
                    sm_id = new_ulid()
                    balance_before = current_physical_qty[sku_code]
                    balance_after = balance_before + order_qty
                    current_physical_qty[sku_code] = balance_after

                    cur.execute(
                        """
                        INSERT INTO inventory_stockmovement (
                            stock_movement_id, company_id, product_variant_id, warehouse_id,
                            movement_type, field_change, quantity,
                            reference_number, note,
                            balance_before, balance_after,
                            cdate, udate
                        ) VALUES (
                            %s, %s, %s, %s,
                            'IN', 'physical_qty', %s,
                            %s, %s,
                            %s, %s,
                            %s, %s
                        )
                        """,
                        (
                            sm_id,
                            company_id,
                            variant_id,
                            warehouse_id,
                            order_qty,
                            po_number,
                            f"PO import: {sheet_name}",
                            balance_before,
                            balance_after,
                            ts,
                            ts,
                        ),
                    )

                    # UPDATE ProductVariantWarehouse — only physical_qty for COMPLETED POs
                    # (incoming_qty tracks in-transit stock; COMPLETED means already received)
                    cur.execute(
                        "UPDATE inventory_productvariantwarehouse "
                        "SET physical_qty = physical_qty + %s "
                        "WHERE product_variant_id = %s AND warehouse_id = %s",
                        (order_qty, variant_id, warehouse_id),
                    )

                    # UPDATE ProductVariant — only total_available_qty for COMPLETED POs
                    cur.execute(
                        "UPDATE inventory_productvariant "
                        "SET total_available_qty = total_available_qty + %s "
                        "WHERE product_variant_id = %s",
                        (order_qty, variant_id),
                    )

                    po_total_qty += order_qty
                    po_total_idr += unit_price_base * order_qty
                    n_rows += 1

                # UPDATE PurchaseOrder totals
                cur.execute(
                    "UPDATE purchasing_purchaseorder "
                    "SET total_ordered_qty = %s, total_received_qty = %s, "
                    "    total_item_amount = %s, procure_amount = %s "
                    "WHERE purchase_order_id = %s",
                    (po_total_qty, po_total_qty, po_total_idr, po_total_idr, po_id),
                )

                print(
                    f"  Sheet: {sheet_name} | PO: {po_number} | Date: {po_date} | Rate: {exchange_rate_int} | Rows: {n_rows} | Skipped: {n_skipped}"
                )

            # Re-enable PO number trigger
            print("\nRe-enabling trigger trg_generate_po_number ...")
            cur.execute(
                "ALTER TABLE purchasing_purchaseorder ENABLE TRIGGER trg_generate_po_number"
            )

            # ------------------------------------------------------------------
            # Step 3 — Final validation queries
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("VALIDATION")
            print("=" * 60)

            # Count only HIST- POs imported by this script
            cur.execute(
                "SELECT COUNT(*) FROM purchasing_purchaseorder "
                "WHERE company_id = %s AND purchase_order_number LIKE 'HIST-%%'",
                (company_id,),
            )
            po_count = cur.fetchone()[0]
            print(f"  HIST PurchaseOrders: {po_count}")

            cur.execute(
                "SELECT COUNT(*) FROM purchasing_purchaseorderdetail "
                "WHERE company_id = %s AND purchase_order_id IN ("
                "  SELECT purchase_order_id FROM purchasing_purchaseorder "
                "  WHERE purchase_order_number LIKE 'HIST-%%')",
                (company_id,),
            )
            pod_count = cur.fetchone()[0]
            print(f"  HIST PO Details: {pod_count}")

            cur.execute(
                "SELECT COUNT(*) FROM inventory_productcogs "
                "WHERE company_id = %s AND reference_number LIKE 'HIST-%%'",
                (company_id,),
            )
            cogs_count = cur.fetchone()[0]
            print(f"  HIST ProductCogs (FIFO): {cogs_count}")

            cur.execute(
                "SELECT COUNT(*) FROM inventory_stockmovement "
                "WHERE company_id = %s AND reference_number LIKE 'HIST-%%'",
                (company_id,),
            )
            sm_count = cur.fetchone()[0]
            print(f"  HIST StockMovements: {sm_count}")

            cur.execute(
                "SELECT SUM(physical_qty) FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s",
                (company_id,),
            )
            total_stock = cur.fetchone()[0] or 0
            print(f"  Total stock (sum of physical_qty): {total_stock}")

            cur.execute(
                "SELECT COUNT(*) FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s AND pvw.physical_qty < 0",
                (company_id,),
            )
            neg_stock = cur.fetchone()[0]
            print(f"  Negative stock rows: {neg_stock}")

            # Additional validation: balance chain integrity
            cur.execute(
                "SELECT COUNT(*) FROM inventory_stockmovement "
                "WHERE balance_after != balance_before + quantity"
            )
            balance_issues = cur.fetchone()[0]
            print(f"  StockMovement balance chain issues: {balance_issues}")

            # Validation checks
            assert po_count == 14, f"Expected 14 HIST POs, got {po_count}"
            assert pod_count == cogs_count == sm_count, (
                f"Mismatch: Details={pod_count}, Cogs={cogs_count}, Movements={sm_count}"
            )
            assert neg_stock == 0, f"Negative stock found: {neg_stock}"
            assert balance_issues == 0, f"Balance chain issues: {balance_issues}"

            print("\n" + "=" * 60)
            print("ALL VALIDATIONS PASSED")
            print("=" * 60)


if __name__ == "__main__":
    main()
