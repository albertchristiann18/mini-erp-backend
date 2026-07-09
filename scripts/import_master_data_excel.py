"""
Phase 2 — Import product catalog (Master SKU + variants) from FinOps Excel.

Run from mini-erp-backend/:
    uv run python scripts/import_master_data_excel.py
"""

import json
from collections import defaultdict
from datetime import datetime, timezone

import openpyxl
import psycopg
import ulid as ulid_lib

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = "/Users/jtf01644/personal/mini-erp-project/mirako_data/Mirakokids FinOps.xlsx"

DB_PARAMS = dict(
    host="localhost",
    port=5433,
    dbname="mini_erp",
    user="postgres",
    password="postgres",
)

# Category code → (name, shopee_category_id)
CATEGORY_META = {
    "SEG": ("Set / Setelan", 525),
    "DRS": ("Dress", 516),
    "JNG": ("Jeans / Pants", 517),
    "PNG": ("Pants", 515),
}

DEFAULT_SHIPPING_CONFIG = {
    "insurance": {"is_required": False, "fee_type": "percentage"},
    "general": {
        "reguler": {
            "cod": True,
            "expeditions": [
                {"code": "anteraja_reguler", "name": "Anteraja Reguler", "is_active": False},
                {"code": "id_express", "name": "ID Express", "is_active": False},
                {"code": "jne", "name": "JNE Reguler", "is_active": False},
                {"code": "ninja_xpress", "name": "Ninja Xpress", "is_active": False},
                {"code": "pos_reguler", "name": "Pos Reguler", "is_active": False},
                {"code": "sicepat", "name": "SiCepat REG", "is_active": False},
                {"code": "jnt", "name": "J&T Express", "is_active": False},
                {"code": "express", "name": "Express", "is_active": False},
            ],
        },
        "instant": {
            "cod": False,
            "expeditions": [
                {"code": "grab", "name": "GrabExpress", "is_active": False},
                {"code": "gojek", "name": "GoSend Instant", "is_active": False},
            ],
        },
        "instant_priority": {
            "cod": False,
            "expeditions": [
                {"code": "grab", "name": "GrabExpress Instant Prioritas", "is_active": False},
                {"code": "gojek", "name": "GoSend Instant Prioritas", "is_active": False},
            ],
        },
        "cargo": {
            "cod": False,
            "expeditions": [
                {"code": "anteraja_cargo", "name": "Anteraja Cargo", "is_active": False},
                {"code": "anteraja_economy", "name": "Anteraja Economy", "is_active": False},
                {"code": "jnt", "name": "J&T Cargo", "is_active": False},
                {"code": "jne", "name": "JNE Trucking (JTR)", "is_active": False},
                {"code": "sentral_cargo", "name": "Sentral Cargo", "is_active": False},
                {"code": "sicepat_gokil", "name": "Sicepat Gokil", "is_active": False},
                {"code": "sicepat_halu", "name": "SiCepat Halu", "is_active": False},
                {"code": "express_eco", "name": "Express Eco", "is_active": False},
            ],
        },
        "sameday": {
            "cod": False,
            "expeditions": [
                {"code": "anteraja", "name": "Anteraja Sameday", "is_active": False},
                {"code": "grab", "name": "GrabExpress Sameday", "is_active": False},
                {"code": "gojek", "name": "GoSend Same Day", "is_active": False},
            ],
        },
        "nextday": {
            "cod": False,
            "expeditions": [
                {"code": "jne", "name": "JNE YES", "is_active": False},
                {"code": "sicepat", "name": "Sicepat BEST", "is_active": False},
            ],
        },
    },
    "marketplaces": {
        "Shopee": {
            "reguler": {
                "expeditions": [
                    {"code": "spx_standard", "name": "SPX Standard", "is_active": False}
                ]
            },
            "cargo": {
                "expeditions": [{"code": "spx_hemat", "name": "SPX Hemat", "is_active": False}]
            },
            "instant": {
                "expeditions": [{"code": "spx_instant", "name": "SPX Instant", "is_active": False}]
            },
            "instant_priority": {
                "expeditions": [
                    {
                        "code": "spx_instant_prio",
                        "name": "SPX Instant Prioritas",
                        "is_active": False,
                    }
                ]
            },
            "sameday": {
                "expeditions": [{"code": "spx_sameday", "name": "SPX Sameday", "is_active": False}]
            },
        },
        "Tokopedia_TikTok": {"use_general_config": True},
    },
}


def new_ulid() -> str:
    """Return a ULID as UUID string (the format PostgreSQL uuid columns expect)."""
    return str(ulid_lib.new().uuid)


def now() -> datetime:
    return datetime.now(timezone.utc)


# ---------------------------------------------------------------------------
# Step 1 — Read Excel sheets
# ---------------------------------------------------------------------------
def load_excel():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    master_sku_rows = list(wb["Master SKU"].iter_rows(values_only=True))
    variant_rows = list(wb["Master SKU Variant"].iter_rows(values_only=True))
    print(f"  Master SKU rows (incl. header): {len(master_sku_rows)}")
    print(f"  Master SKU Variant rows (incl. header): {len(variant_rows)}")
    return master_sku_rows, variant_rows


# ---------------------------------------------------------------------------
# Step 2 — Parse variant sheet and build product→variants map
# ---------------------------------------------------------------------------
def parse_variants(variant_rows):
    """
    Returns:
        variants_by_sku: dict[sku_code → list of variant dicts]
        all_supplier_names: set of supplier names from variant sheet
    """
    variants_by_sku: dict[str, list] = defaultdict(list)
    all_supplier_names: set[str] = set()

    skipped = 0
    for row in variant_rows[1:]:  # skip header
        sku_variant_code = row[6]  # col 7 (0-indexed: 6)
        if not sku_variant_code or str(sku_variant_code).strip() == "":
            skipped += 1
            continue

        sku_code = row[5]  # col 6
        if not sku_code or str(sku_code).strip() == "":
            skipped += 1
            continue

        sku_code = str(sku_code).strip()
        sku_variant_code = str(sku_variant_code).strip()

        color_code = str(row[2]).strip().lower() if row[2] else ""  # col 3
        product_name = str(row[7]).strip() if row[7] else ""  # col 8
        supplier_name = str(row[8]).strip() if row[8] else ""  # col 9
        color_display = str(row[9]).strip() if row[9] else ""  # col 10
        # col 12 = Order Size
        order_size_raw = row[12]
        size_code = (
            str(int(order_size_raw))
            if isinstance(order_size_raw, float)
            else str(order_size_raw).strip()
            if order_size_raw
            else ""
        )
        # col 13 = COGS
        cogs_raw = row[13]
        cogs = int(cogs_raw) if cogs_raw is not None else 0
        # col 19 = Selling Price (0-indexed)
        price_raw = row[19]
        base_price = int(price_raw) if price_raw is not None else 0
        # col 27 = Stock Master Warehouse (0-indexed)
        stock_raw = row[27]
        stock_qty = int(stock_raw) if stock_raw is not None else 0

        if supplier_name:
            all_supplier_names.add(supplier_name)

        variants_by_sku[sku_code].append(
            {
                "sku_variant_code": sku_variant_code,
                "product_name": product_name,
                "supplier_name": supplier_name,
                "color_code": color_code,
                "color_display": color_display,
                "size_code": size_code,
                "cogs": cogs,
                "base_price": base_price,
                "stock_qty": stock_qty,
            }
        )

    print(f"  Variant rows skipped (empty SKU VARIANT CODE): {skipped}")
    return variants_by_sku, all_supplier_names


# ---------------------------------------------------------------------------
# Step 3 — Build variant_options JSON per product
# ---------------------------------------------------------------------------
def build_variant_options(variant_list):
    """Build variant_options, dim1_options, dim2_options for a product."""
    # dim1 = size, dim2 = color
    sizes_seen = []
    size_set = set()
    color_map = {}  # color_code → color_display

    for v in variant_list:
        if v["size_code"] and v["size_code"] not in size_set:
            sizes_seen.append(v["size_code"])
            size_set.add(v["size_code"])
        if v["color_code"] and v["color_code"] not in color_map:
            color_map[v["color_code"]] = v["color_display"]

    # Sort sizes numerically if possible
    def size_sort_key(s):
        try:
            return (0, int(s))
        except (ValueError, TypeError):
            return (1, s)

    dim1_options = sorted(sizes_seen, key=size_sort_key)
    dim2_options = sorted(color_map.keys())

    variant_options = [
        {
            "id": "size",
            "name": "Size",
            "order": 1,
            "values": [{"id": s, "label": s} for s in dim1_options],
        },
        {
            "id": "color",
            "name": "Color",
            "order": 2,
            "values": [{"id": cc, "label": color_map[cc]} for cc in dim2_options],
        },
    ]
    return variant_options, dim1_options, dim2_options


# ---------------------------------------------------------------------------
# Main import
# ---------------------------------------------------------------------------
def main():
    ts = now()

    # --- Load Excel ---
    master_sku_rows, variant_rows = load_excel()

    # --- Parse variants ---
    variants_by_sku, variant_supplier_names = parse_variants(variant_rows)

    # --- Collect supplier names from Master SKU sheet (col 5, 0-indexed 4) ---
    master_sku_supplier_names: set[str] = set()
    for row in master_sku_rows[1:]:
        supplier_raw = row[4]  # col 5
        if supplier_raw and str(supplier_raw).strip():
            master_sku_supplier_names.add(str(supplier_raw).strip())

    all_supplier_names = master_sku_supplier_names | variant_supplier_names
    print(f"Suppliers found: {sorted(all_supplier_names)}")

    # --- Collect category codes from Master SKU (col 6, 0-indexed 5) ---
    category_codes_found: set[str] = set()
    for row in master_sku_rows[1:]:
        cat_code = row[5]  # col 6
        if cat_code and str(cat_code).strip():
            category_codes_found.add(str(cat_code).strip())
    print(f"Category codes found: {sorted(category_codes_found)}")

    # --- Connect to DB ---
    # autocommit=True is required so ALTER TABLE DISABLE/ENABLE TRIGGER
    # can run outside any transaction (PostgreSQL restriction).
    # All inserts use ON CONFLICT DO NOTHING for idempotency.
    print("\nConnecting to database...")
    with psycopg.connect(**DB_PARAMS, autocommit=True) as conn:
        with conn.cursor() as cur:
            # ---------------------------------------------------------------
            # Step 1 — Discover company (autocommit connection for DDL)
            # ---------------------------------------------------------------
            cur.execute("SELECT company_id FROM core_company WHERE name = 'Mirako' LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise RuntimeError(
                    "Company 'Mirako' not found — run: "
                    "uv run python scripts/create_user.py --username admin --role admin "
                    "--password admin123 --company Mirako"
                )
            company_id = str(row[0])
            print(f"Company 'Mirako' id: {company_id}")

            # ---------------------------------------------------------------
            # Step 2 — Ensure warehouse exists
            # ---------------------------------------------------------------
            cur.execute(
                "SELECT warehouse_id FROM master_warehouse WHERE name = 'Master Warehouse' AND company_id = %s LIMIT 1",
                (company_id,),
            )
            row = cur.fetchone()
            if row:
                warehouse_id = str(row[0])
                print(f"Warehouse 'Master Warehouse' already exists: {warehouse_id}")
            else:
                warehouse_id = new_ulid()
                cur.execute(
                    """
                    INSERT INTO master_warehouse
                      (warehouse_id, company_id, name, address, is_marketplace_visible, is_active, cdate, udate)
                    VALUES (%s, %s, 'Master Warehouse', NULL, TRUE, TRUE, %s, %s)
                    ON CONFLICT DO NOTHING
                    """,
                    (warehouse_id, company_id, ts, ts),
                )
                print(f"Created warehouse 'Master Warehouse': {warehouse_id}")

            # ---------------------------------------------------------------
            # Step 3 — Create suppliers
            # ---------------------------------------------------------------
            supplier_map: dict[str, str] = {}  # name → supplier_id
            suppliers_created = 0
            for supplier_name in sorted(all_supplier_names):
                # Check if already exists
                cur.execute(
                    "SELECT supplier_id FROM inventory_supplier WHERE name = %s AND company_id = %s LIMIT 1",
                    (supplier_name, company_id),
                )
                existing = cur.fetchone()
                if existing:
                    supplier_map[supplier_name] = str(existing[0])
                else:
                    sid = new_ulid()
                    cur.execute(
                        """
                        INSERT INTO inventory_supplier
                          (supplier_id, company_id, name, is_active, cdate, udate)
                        VALUES (%s, %s, %s, TRUE, %s, %s)
                        ON CONFLICT (supplier_id) DO NOTHING
                        """,
                        (sid, company_id, supplier_name, ts, ts),
                    )
                    supplier_map[supplier_name] = sid
                    suppliers_created += 1
                    print(f"  Created supplier: {supplier_name} ({sid})")

            # ---------------------------------------------------------------
            # Step 4 — Create categories
            # category_code is globally unique in this schema — reuse existing
            # rows from any company (category codes like SEG, DRS are shared).
            # If a category doesn't exist at all, create it under Mirako.
            # ---------------------------------------------------------------
            cat_map: dict[str, str] = {}  # category_code → category_id
            categories_inserted = 0
            for cat_code in sorted(category_codes_found):
                meta = CATEGORY_META.get(cat_code)
                if not meta:
                    print(f"  WARNING: Unknown category code '{cat_code}' — skipping")
                    continue
                cat_name, shopee_cat_id = meta

                # Look up by category_code (globally unique in this schema)
                cur.execute(
                    "SELECT category_id FROM master_category WHERE category_code = %s LIMIT 1",
                    (cat_code,),
                )
                existing = cur.fetchone()
                if existing:
                    cat_map[cat_code] = str(existing[0])
                    print(f"  Category already exists: {cat_code} ({cat_map[cat_code]})")
                else:
                    cid = new_ulid()
                    cur.execute(
                        """
                        INSERT INTO master_category
                          (category_id, company_id, name, category_code,
                           shopee_category_id, is_active, master_category_key, cdate, udate)
                        VALUES (%s, %s, %s, %s, %s, TRUE, '', %s, %s)
                        ON CONFLICT (category_code) DO NOTHING
                        """,
                        (cid, company_id, cat_name, cat_code, shopee_cat_id, ts, ts),
                    )
                    # Re-fetch after insert
                    cur.execute(
                        "SELECT category_id FROM master_category WHERE category_code = %s LIMIT 1",
                        (cat_code,),
                    )
                    cat_map[cat_code] = str(cur.fetchone()[0])
                    categories_inserted += 1
                    print(f"  Created category: {cat_code} → {cat_name}")

            print(f"\nCategories inserted: {categories_inserted} (total in map: {len(cat_map)})")

            # ---------------------------------------------------------------
            # Step 5 — Disable SKU trigger, insert products, re-enable
            # ---------------------------------------------------------------
            print("\nDisabling trigger trg_generate_sku ...")
            cur.execute("ALTER TABLE master_product DISABLE TRIGGER trg_generate_sku")

            products_inserted = 0
            products_skipped = 0
            product_id_map: dict[str, str] = {}  # sku_code → product_id
            product_supplier_map: dict[str, str] = {}  # sku_code → supplier_name

            for row in master_sku_rows[1:]:  # skip header
                sku_code_raw = row[7]  # col 8
                if not sku_code_raw or str(sku_code_raw).strip() == "":
                    products_skipped += 1
                    continue

                sku_code = str(sku_code_raw).strip()
                cat_code = str(row[5]).strip() if row[5] else ""  # col 6
                product_name = str(row[9]).strip() if row[9] else ""  # col 10
                supplier_name = str(row[4]).strip() if row[4] else ""  # col 5

                if cat_code not in cat_map:
                    print(f"  SKIP product {sku_code}: unknown category '{cat_code}'")
                    products_skipped += 1
                    continue

                category_id = cat_map[cat_code]

                # Build variant options from variants_by_sku
                variant_list = variants_by_sku.get(sku_code, [])
                variant_options, dim1_options, dim2_options = build_variant_options(variant_list)

                pid = new_ulid()

                cur.execute(
                    """
                    INSERT INTO master_product (
                      product_id, company_id, category_id, name, sku_code,
                      total_qty, total_cogs,
                      variant_options, specifications, shipping_config,
                      weight, length, width, height,
                      dim1_key, dim2_key, dim1_options, dim2_options,
                      is_active, cdate, udate
                    ) VALUES (
                      %s, %s, %s, %s, %s,
                      0, 0,
                      %s, %s, %s,
                      0, 0, 0, 0,
                      'size', 'color', %s, %s,
                      TRUE, %s, %s
                    )
                    ON CONFLICT (sku_code) DO NOTHING
                    """,
                    (
                        pid,
                        company_id,
                        category_id,
                        product_name,
                        sku_code,
                        json.dumps(variant_options),
                        json.dumps({}),
                        json.dumps(DEFAULT_SHIPPING_CONFIG),
                        json.dumps(dim1_options),
                        json.dumps(dim2_options),
                        ts,
                        ts,
                    ),
                )

                # Fetch actual product_id (may already exist via ON CONFLICT)
                cur.execute(
                    "SELECT product_id FROM master_product WHERE sku_code = %s LIMIT 1",
                    (sku_code,),
                )
                actual_pid = str(cur.fetchone()[0])
                product_id_map[sku_code] = actual_pid

                if cur.rowcount > 0 or True:
                    products_inserted += 1

                if supplier_name:
                    product_supplier_map[sku_code] = supplier_name

            print("Re-enabling trigger trg_generate_sku ...")
            cur.execute("ALTER TABLE master_product ENABLE TRIGGER trg_generate_sku")

            # Reset sequence
            cur.execute(
                "SELECT setval('product_sku_seq', GREATEST((SELECT COUNT(*) FROM master_product WHERE company_id = %s), 1))",
                (company_id,),
            )
            seq_val = cur.fetchone()[0]
            print(f"Sequence reset to: {seq_val}")

            print(f"\nProducts processed: {products_inserted}, skipped: {products_skipped}")

            # ---------------------------------------------------------------
            # Step 5b — Insert ProductSupplier links
            # ---------------------------------------------------------------
            ps_inserted = 0
            for sku_code, supplier_name in product_supplier_map.items():
                if sku_code not in product_id_map:
                    continue
                if supplier_name not in supplier_map:
                    continue
                product_id = product_id_map[sku_code]
                supplier_id = supplier_map[supplier_name]
                ps_id = new_ulid()
                cur.execute(
                    """
                    INSERT INTO inventory_productsupplier
                      (product_supplier_id, company_id, product_id, supplier_id, supplier_link, cdate, udate)
                    VALUES (%s, %s, %s, %s, NULL, %s, %s)
                    ON CONFLICT (product_id, supplier_id) DO NOTHING
                    """,
                    (ps_id, company_id, product_id, supplier_id, ts, ts),
                )
                ps_inserted += 1

            print(f"ProductSupplier rows: {ps_inserted}")

            # ---------------------------------------------------------------
            # Step 6 — Insert variants and warehouse stock
            # ---------------------------------------------------------------
            variants_inserted = 0
            wh_stock_inserted = 0
            variants_skipped = 0

            for row in variant_rows[1:]:  # skip header
                sku_variant_code = row[6]  # col 7
                if not sku_variant_code or str(sku_variant_code).strip() == "":
                    variants_skipped += 1
                    continue

                sku_code_raw = row[5]  # col 6
                if not sku_code_raw or str(sku_code_raw).strip() == "":
                    variants_skipped += 1
                    continue

                sku_code = str(sku_code_raw).strip()
                sku_variant_code = str(sku_variant_code).strip()

                if sku_code not in product_id_map:
                    print(
                        f"  SKIP variant {sku_variant_code}: parent product '{sku_code}' not found"
                    )
                    variants_skipped += 1
                    continue

                product_id = product_id_map[sku_code]

                color_code = str(row[2]).strip().lower() if row[2] else ""  # col 3
                product_name = str(row[7]).strip() if row[7] else ""  # col 8
                color_display = str(row[9]).strip() if row[9] else ""  # col 10

                order_size_raw = row[12]  # col 13
                size_code = (
                    str(int(order_size_raw))
                    if isinstance(order_size_raw, float)
                    else str(order_size_raw).strip()
                    if order_size_raw
                    else ""
                )

                cogs_raw = row[13]  # col 14
                cogs = int(cogs_raw) if cogs_raw is not None else 0

                price_raw = row[19]  # col 20
                base_price = int(price_raw) if price_raw is not None else 0

                stock_raw = row[27]  # col 28
                stock_qty = int(stock_raw) if stock_raw is not None else 0

                variant_name = f"{size_code} / {color_display}"
                variant_values = {"size": size_code, "color": color_code}

                vid = new_ulid()
                cur.execute(
                    """
                    INSERT INTO master_productvariant (
                      product_variant_id, company_id, product_id, name,
                      sku_variant_code, variant_values,
                      current_cogs, base_price,
                      total_incoming_qty, total_outgoing_qty, total_available_qty,
                      is_active, is_fake, cdate, udate
                    ) VALUES (
                      %s, %s, %s, %s,
                      %s, %s,
                      %s, %s,
                      %s, 0, %s,
                      TRUE, FALSE, %s, %s
                    )
                    ON CONFLICT (sku_variant_code) DO NOTHING
                    """,
                    (
                        vid,
                        company_id,
                        product_id,
                        variant_name,
                        sku_variant_code,
                        json.dumps(variant_values),
                        cogs,
                        base_price,
                        stock_qty,
                        stock_qty,
                        ts,
                        ts,
                    ),
                )
                variants_inserted += 1

                # Fetch actual variant id
                cur.execute(
                    "SELECT product_variant_id FROM master_productvariant WHERE sku_variant_code = %s LIMIT 1",
                    (sku_variant_code,),
                )
                actual_vid = str(cur.fetchone()[0])

                # Warehouse stock
                pvw_id = new_ulid()
                cur.execute(
                    """
                    INSERT INTO inventory_productvariantwarehouse (
                      product_variant_warehouse_id, company_id, product_variant_id, warehouse_id,
                      incoming_qty, outgoing_qty, physical_qty, checkout_qty,
                      cdate, udate
                    ) VALUES (
                      %s, %s, %s, %s,
                      %s, 0, %s, 0,
                      %s, %s
                    )
                    ON CONFLICT (product_variant_id, warehouse_id) DO NOTHING
                    """,
                    (
                        pvw_id,
                        company_id,
                        actual_vid,
                        warehouse_id,
                        stock_qty,
                        stock_qty,
                        ts,
                        ts,
                    ),
                )
                wh_stock_inserted += 1

            print(f"Variants processed: {variants_inserted}, skipped: {variants_skipped}")
            print(f"Warehouse stock rows: {wh_stock_inserted}")

            # ---------------------------------------------------------------
            # Step 7 — Summary (autocommit: no explicit commit needed)
            # ---------------------------------------------------------------
            cur.execute("SELECT COUNT(*) FROM master_category WHERE company_id = %s", (company_id,))
            n_cats = cur.fetchone()[0]

            cur.execute("SELECT COUNT(*) FROM master_product WHERE company_id = %s", (company_id,))
            n_products = cur.fetchone()[0]

            cur.execute(
                "SELECT COUNT(*) FROM master_productvariant WHERE company_id = %s", (company_id,)
            )
            n_variants = cur.fetchone()[0]

            cur.execute(
                "SELECT COUNT(*) FROM inventory_productvariantwarehouse pvw "
                "JOIN master_productvariant pv ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s",
                (company_id,),
            )
            n_stocks = cur.fetchone()[0]

            cur.execute(
                "SELECT COUNT(*) FROM master_productvariant WHERE sku_variant_code = '' AND company_id = %s",
                (company_id,),
            )
            n_empty_codes = cur.fetchone()[0]

        print("\n" + "=" * 60)
        print("IMPORT SUMMARY")
        print("=" * 60)
        print(f"  Categories:           {n_cats}")
        print(f"  Products:             {n_products}")
        print(f"  Variants:             {n_variants}")
        print(f"  Warehouse stock rows: {n_stocks}")
        print(f"  Empty sku_variant_codes: {n_empty_codes} (should be 0)")
        print("=" * 60)


if __name__ == "__main__":
    main()
