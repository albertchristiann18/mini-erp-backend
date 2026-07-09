"""
Mirako Data Migration — Master Script

Migrates a blank database to fully seeded state in one command.
Source data: mirako_data/ folder (sibling of mini-erp-backend/).

Usage:
    uv run python scripts/migrate_all_mirako_kids.py
    uv run python scripts/migrate_all_mirako_kids.py --skip-files   # skip R2 uploads

Steps:
    1. Reset      — truncate all business tables
    2. Seed       — create company Mirako + superuser admin/admin123
    3. Master     — products, variants, categories, suppliers
    3b. Links     — 1688 supplier URLs per product (from PO Detail Overview)
    4. PO         — purchase orders + line items + FIFO
    5. PO enrich  — invoice numbers, delivery orders, fees from PO Tracker
    5b. Split     — create split-delivery POs (PO-2025-001B/002B, PO-2026-001B)
    6. Sales      — sales orders from Sales Order Tracker
    7. Cash       — cash transactions from FinOps Finance sheet
    8. Files      — upload PO invoices, resi, DO invoices to R2
    9. Report     — print reconciliation counts
"""

import argparse
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

# ---------------------------------------------------------------------------
# Bootstrap Django before any Django imports
# ---------------------------------------------------------------------------
sys.path.insert(0, str(Path(__file__).parent.parent))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
import django

django.setup()

import openpyxl
import psycopg
import ulid as ulid_lib
from django.contrib.auth import get_user_model
from django.core.files import File
from django.core.files.storage import default_storage

from core.models import Company, UserProfile

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MIRAKO_DATA = Path(__file__).parent.parent.parent / "mirako_data"
OPEX_DOCS = MIRAKO_DATA / "opex docs"

DB_PARAMS = dict(
    host="localhost",
    port=5433,
    dbname="mini_erp",
    user="postgres",
    password="postgres",
)

SCRIPTS_DIR = Path(__file__).parent

IMPORTED_PO_NUMBERS = [f"PO-2025-{i:03d}" for i in range(1, 8)] + [
    f"PO-2026-{i:03d}" for i in range(1, 8)
]

# tracker "no" (1-14) -> PO number
TRACKER_NO_TO_PO = {
    1: "PO-2025-001",
    2: "PO-2025-002",
    3: "PO-2025-003",
    4: "PO-2025-004",
    5: "PO-2025-005",
    6: "PO-2025-006",
    7: "PO-2025-007",
    8: "PO-2026-001",
    9: "PO-2026-002",
    10: "PO-2026-003",
    11: "PO-2026-004",
    12: "PO-2026-005",
    13: "PO-2026-006",
    14: "PO-2026-007",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def header(title: str) -> None:
    print(f"\n{'=' * 60}")
    print(f"  {title}")
    print(f"{'=' * 60}")


def run_script(name: str, script: str) -> None:
    """Run a migration script as subprocess; exit on failure."""
    header(f"Step: {name}")
    result = subprocess.run(
        ["uv", "run", "python", str(SCRIPTS_DIR / script)],
        cwd=str(SCRIPTS_DIR.parent),
    )
    if result.returncode != 0:
        print(f"\n[FAIL] {name} exited with code {result.returncode}")
        sys.exit(1)
    print(f"\n[OK] {name}")


def _clean_resi_filename(raw: str) -> str:
    """'RESI-GZ3-41906.pdf (WZ)' -> 'RESI-GZ3-41906.pdf'"""
    return re.sub(r"\s*\([^)]*\)\s*$", "", raw.strip())


def _clean_doi_filename(raw: str) -> str:
    """'INVOICE/BRG/81524.pdf' -> 'INVOICE_BRG_81524.pdf'"""
    return raw.replace("/", "_")


def _to_date(v) -> date | None:
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------------
# Step 1 — Reset
# ---------------------------------------------------------------------------
def step_reset() -> None:
    header("Step 1: Reset business tables")
    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()
    cur.execute("""
        TRUNCATE TABLE
            sales_salesreturnitem, sales_salesreturn,
            sales_salesordercogsdetail, sales_salesorderitem, sales_salesorder,
            inventory_productcogs,
            finance_paymentrecord, finance_accountspayable,
            finance_accountsreceivable, finance_cashtransaction,
            purchasing_purchaseorderstatushistory,
            purchasing_purchaseorderdetail, purchasing_purchaseorder,
            po_number_counter,
            inventory_stockmovement,
            inventory_productvariantwarehouse, master_productvariantmarketplace,
            inventory_productbusinessentity, inventory_productsupplier,
            inventory_productdimensionimage, master_productphoto,
            purchasing_sourcingpoolitem, purchasing_sourcingpool,
            master_productvariant, master_product,
            inventory_supplier, master_category
        RESTART IDENTITY CASCADE
    """)
    conn.close()
    print("[OK] All business tables truncated")


# ---------------------------------------------------------------------------
# Step 2 — Seed user & company
# ---------------------------------------------------------------------------
def step_seed_user() -> None:
    header("Step 2: Seed user & company")
    User = get_user_model()

    company, created = Company.objects.get_or_create(
        name="Mirako",
        defaults={"is_active": True},
    )
    print(f"{'Created' if created else 'Existing'} company: {company.name} (id={company.id})")

    USERNAME = "admin"
    PASSWORD = "admin123"

    if not User.objects.filter(username=USERNAME).exists():
        user = User.objects.create_superuser(
            username=USERNAME,
            email="albert.christiann18@gmail.com",
            password=PASSWORD,
        )
        user_created = True
    else:
        user = User.objects.get(username=USERNAME)
        user_created = False

    print(f"{'Created' if user_created else 'Existing'} user: {USERNAME} / {PASSWORD}")

    profile, profile_created = UserProfile.objects.get_or_create(
        user=user,
        defaults={"company": company, "role": "admin"},
    )
    if not profile_created and profile.company != company:
        profile.company = company
        profile.save(update_fields=["company"])

    print(f"Profile role={profile.role}, company={profile.company.name}")
    print(f"\n  Login: {USERNAME} / {PASSWORD}")


# ---------------------------------------------------------------------------
# Step 3 — Master data (products, variants, categories, suppliers)
# ---------------------------------------------------------------------------
def step_master_data() -> None:
    run_script("Master Data", "import_master_data_excel.py")
    # Apply default dimensions to all products (weight in grams, dimensions in cm)
    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    conn.cursor().execute(
        "UPDATE master_product SET weight = 200, height = 1, width = 30, length = 30"
    )
    conn.close()
    print("[OK] Default dimensions set on all products (200g, 1×30×30 cm)")


# ---------------------------------------------------------------------------
# Step 3b — Product supplier links (1688 URLs)
# ---------------------------------------------------------------------------
def step_product_links() -> None:
    header("Step 3b: Product supplier links (1688 URLs)")
    po_detail_path = MIRAKO_DATA / "Purchase Order Detail Overview (1).xlsx"
    wb = openpyxl.load_workbook(po_detail_path, data_only=True)
    ws = wb["Master DATA All Item"]
    rows = list(ws.iter_rows(values_only=True))
    # Header at row index 1, data from row 2
    # col 9: variant SKU CODE (e.g. "DRS-008-100-BLU")
    # col 1: 1688 Product Link

    # Build map: product SKU prefix (first two dash-parts) -> first non-null link
    link_map: dict[str, str] = {}
    for row in rows[2:]:
        sku_code = row[9]
        link = row[1]
        if not sku_code or not link:
            continue
        sku_str = str(sku_code).strip()
        link_str = str(link).strip()
        if not link_str.startswith("http"):
            continue
        parts = sku_str.split("-")
        if len(parts) < 2:
            continue
        product_prefix = f"{parts[0]}-{parts[1]}"
        if product_prefix not in link_map:
            link_map[product_prefix] = link_str

    print(f"  Found {len(link_map)} product links in Excel")

    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()

    updated = 0
    for sku_code, link in link_map.items():
        cur.execute(
            """
            UPDATE inventory_productsupplier ps
            SET supplier_link = %s
            FROM master_product p
            WHERE ps.product_id = p.product_id
              AND p.sku_code = %s
            """,
            (link, sku_code),
        )
        if cur.rowcount:
            print(f"  ✓ {sku_code}: {link[:70]}")
            updated += cur.rowcount
        else:
            print(f"  SKIP {sku_code}: no product-supplier record found")

    conn.close()
    print(f"\n[OK] Updated supplier_link on {updated} product-supplier records")


# ---------------------------------------------------------------------------
# Step 4 — Purchase orders
# ---------------------------------------------------------------------------
def step_purchase_orders() -> None:
    run_script("Purchase Orders", "import_purchase_orders_excel.py")


# ---------------------------------------------------------------------------
# Step 5 — PO tracker enrichment
# ---------------------------------------------------------------------------
def parse_po_tracker() -> dict:
    """
    Read Purchase Order Tracker (1).xlsx and return per-PO metadata dict.
    Keys are PO numbers (e.g. "PO-2025-001").
    """
    wb = openpyxl.load_workbook(MIRAKO_DATA / "Purchase Order Tracker (1).xlsx", data_only=True)
    ws = wb["Master Tracker Purchase Order"]
    rows = list(ws.iter_rows(values_only=True))
    # Header at row index 6, data from row index 7
    by_no: dict[int, list] = defaultdict(list)
    for row in rows[7:]:
        no = row[0]
        if not isinstance(no, (int, float)):
            continue
        no_int = int(no)
        if no_int not in TRACKER_NO_TO_PO:
            continue
        by_no[no_int].append(row)

    result = {}
    for no, po_rows in by_no.items():
        first = po_rows[0]
        po_number = TRACKER_NO_TO_PO[no]

        # DO NUMBER cleaning (col 4)
        raw_do = str(first[4]) if first[4] else ""
        resi_filename = _clean_resi_filename(raw_do) if raw_do else ""
        resi_number = re.sub(r"\.pdf.*$", "", resi_filename, flags=re.IGNORECASE).strip()

        # DO invoice filename (col 19)
        raw_doi = str(first[19]) if first[19] else ""
        doi_filename = _clean_doi_filename(raw_doi) if raw_doi else ""

        # PO invoice file (col 12) — only attach if it's a PDF
        po_inv_raw = str(first[12]) if first[12] else ""
        po_invoice_file = po_inv_raw if po_inv_raw.lower().endswith(".pdf") else None

        # Aggregate per-delivery values (sum across all delivery rows for this PO)
        # delivery_fee: col 30 = "shipping 1688" which is already in RMB (China-side fee)
        delivery_fee_rmb = round(float(po_rows[0][30] or 0), 3)
        shipping_fee = int(sum((row[16] or 0) for row in po_rows))
        cbm = round(sum((row[20] or 0) for row in po_rows), 3)
        weight = round(sum((row[23] or 0) for row in po_rows), 3)

        result[po_number] = {
            "invoice_number": str(first[2]) if first[2] else None,
            "invoice_date": _to_date(first[3]),
            "created_date": _to_date(first[1]),
            "delivery_order_number": resi_number or None,
            "delivery_date": _to_date(first[5]),
            "forwarder_name": str(first[9]) if first[9] else None,
            "shop_services": str(first[10]) if first[10] else None,
            "commission_fee_pct": int(round((first[11] or 0) * 100)),
            "delivery_fee_rmb": delivery_fee_rmb,
            "shipping_fee": shipping_fee,
            "cbm": cbm,
            "weight": weight,
            # file attachment
            "po_invoice_file": po_invoice_file,
            "resi_file": resi_filename or None,
            "do_invoice_file": doi_filename or None,
        }

    return result


def step_enrich_po_tracker() -> None:
    header("Step 5: PO Tracker enrichment")
    metadata = parse_po_tracker()

    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()

    updated = 0
    for po_number, m in metadata.items():
        # Fetch current total_item_amount and exchange_rate to recalculate totals
        cur.execute(
            "SELECT total_item_amount, exchange_rate FROM purchasing_purchaseorder "
            "WHERE purchase_order_number = %s",
            (po_number,),
        )
        row = cur.fetchone()
        if not row:
            print(f"  WARNING: PO {po_number} not found — skipping enrichment")
            continue
        total_item_amount = row[0] or 0
        exchange_rate = float(row[1] or 1)

        # delivery_fee is already in RMB (shipping_1688 col); convert to IDR for totals
        delivery_fee_rmb = m["delivery_fee_rmb"]
        delivery_fee_idr_calc = int(round(delivery_fee_rmb * exchange_rate))
        commission_fee = round(m["commission_fee_pct"] / 100 * total_item_amount)
        shipping_fee = m["shipping_fee"]
        total_order_amount = total_item_amount + commission_fee + delivery_fee_idr_calc
        total_amount = total_item_amount + commission_fee + shipping_fee + delivery_fee_idr_calc

        # Build cdate from tracker "created date" col, fall back to invoice_date
        created_ts = None
        if m["created_date"]:
            from datetime import (
                datetime as _dt,
                timezone as _tz,
            )

            created_ts = _dt.combine(m["created_date"], _dt.min.time()).replace(tzinfo=_tz.utc)

        cur.execute(
            """
            UPDATE purchasing_purchaseorder SET
                invoice_number        = %s,
                invoice_date          = %s,
                delivery_order_number = %s,
                delivery_date         = %s,
                forwarder_name        = %s,
                shop_services         = %s,
                commission_fee_pct    = %s,
                delivery_fee          = %s,
                shipping_fee          = %s,
                cbm                   = %s,
                weight                = %s,
                total_order_amount    = %s,
                total_amount          = %s,
                cdate                 = COALESCE(%s, cdate)
            WHERE purchase_order_number = %s
            """,
            (
                m["invoice_number"],
                m["invoice_date"],
                m["delivery_order_number"],
                m["delivery_date"],
                m["forwarder_name"],
                m["shop_services"],
                m["commission_fee_pct"],
                delivery_fee_rmb,
                m["shipping_fee"],
                m["cbm"],
                m["weight"],
                total_order_amount,
                total_amount,
                created_ts,
                po_number,
            ),
        )
        print(f"  ✓ {po_number}: inv={m['invoice_number']}, DO={m['delivery_order_number']}")
        updated += 1

    # Set DRAFT status for POs that are not yet completed
    cur.execute(
        "UPDATE purchasing_purchaseorder SET status = 'DRAFT' "
        "WHERE purchase_order_number IN ('PO-2026-006', 'PO-2026-007')"
    )
    print("  ✓ Set PO-2026-006 and PO-2026-007 to DRAFT")

    conn.close()
    print(f"\n[OK] Enriched {updated} POs from tracker")


# ---------------------------------------------------------------------------
# Step 6 — Sales orders
# ---------------------------------------------------------------------------
def step_sales_orders() -> None:
    run_script("Sales Orders", "import_sales_excel.py")


# ---------------------------------------------------------------------------
# Step 7 — Cash transactions
# ---------------------------------------------------------------------------
def step_cash_transactions() -> None:
    run_script("Cash Transactions", "import_cash_transactions_excel.py")


# ---------------------------------------------------------------------------
# Step 8 — File attachments
# ---------------------------------------------------------------------------
def _upload_file(cur, local_path: Path, storage_key: str, po_number: str, field: str) -> bool:
    """Upload a file to R2 and UPDATE the DB field. Returns True on success."""
    if not local_path.exists():
        print(f"  SKIP  {po_number}.{field}: file not found — {local_path.name}")
        return False
    try:
        if default_storage.exists(storage_key):
            default_storage.delete(storage_key)
        with open(local_path, "rb") as f:
            saved_name = default_storage.save(storage_key, File(f))
        cur.execute(
            f"UPDATE purchasing_purchaseorder SET {field} = %s WHERE purchase_order_number = %s",
            (saved_name, po_number),
        )
        print(f"  ✓ {po_number}.{field} ← {local_path.name}")
        return True
    except Exception as e:
        print(f"  ERROR {po_number}.{field}: {e}")
        return False


def step_attach_files(skip_files: bool) -> None:
    header("Step 8: File attachments")
    if skip_files:
        print("[SKIP] --skip-files flag set")
        return

    metadata = parse_po_tracker()
    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()

    # Clear existing file references on imported POs
    cur.execute(
        """
        UPDATE purchasing_purchaseorder
        SET purchase_order_invoice_file = NULL,
            delivery_order_file = NULL,
            delivery_order_invoice_file = NULL
        WHERE purchase_order_number = ANY(%s)
        """,
        (IMPORTED_PO_NUMBERS,),
    )

    # Also clear split POs
    split_po_numbers = [c["split_po"] for c in SPLIT_DELIVERY_CONFIG]
    cur.execute(
        """
        UPDATE purchasing_purchaseorder
        SET delivery_order_file = NULL, delivery_order_invoice_file = NULL
        WHERE purchase_order_number = ANY(%s)
        """,
        (split_po_numbers,),
    )

    po_inv_count = do_count = doi_count = 0

    for po_number, m in metadata.items():
        # 1. PO invoice file → opex docs/purchase order invoice/
        if m["po_invoice_file"]:
            local = OPEX_DOCS / "purchase order invoice" / m["po_invoice_file"]
            key = "po/invoices/" + m["po_invoice_file"].replace(" ", "_")
            if _upload_file(cur, local, key, po_number, "purchase_order_invoice_file"):
                po_inv_count += 1

        # 2. RESI / delivery order file → opex docs/resi/
        if m["resi_file"]:
            local = OPEX_DOCS / "resi" / m["resi_file"]
            key = "po/delivery_orders/" + m["resi_file"].replace(" ", "_")
            if _upload_file(cur, local, key, po_number, "delivery_order_file"):
                do_count += 1

        # 3. DO invoice file → opex docs/invoice delivery file/
        if m["do_invoice_file"]:
            local = OPEX_DOCS / "invoice delivery file" / m["do_invoice_file"]
            key = "po/do_invoices/" + m["do_invoice_file"].replace(" ", "_")
            if _upload_file(cur, local, key, po_number, "delivery_order_invoice_file"):
                doi_count += 1

    # Attach RESI and DO invoice for split POs (second tracker row per split no)
    wb_t = openpyxl.load_workbook(MIRAKO_DATA / "Purchase Order Tracker (1).xlsx", data_only=True)
    ws_t = wb_t["Master Tracker Purchase Order"]
    t_rows = list(ws_t.iter_rows(values_only=True))
    by_no: dict[int, list] = defaultdict(list)
    for row in t_rows[7:]:
        no = row[0]
        if isinstance(no, (int, float)) and int(no) in TRACKER_NO_TO_PO:
            by_no[int(no)].append(row)

    for cfg in SPLIT_DELIVERY_CONFIG:
        rows_for_no = by_no.get(cfg["tracker_no"], [])
        if len(rows_for_no) < 2:
            continue
        split_row = rows_for_no[1]
        split_po = cfg["split_po"]

        raw_do = str(split_row[4]) if split_row[4] else ""
        resi_fn = _clean_resi_filename(raw_do) if raw_do else ""
        if resi_fn:
            local = OPEX_DOCS / "resi" / resi_fn
            key = "po/delivery_orders/" + resi_fn.replace(" ", "_")
            if _upload_file(cur, local, key, split_po, "delivery_order_file"):
                do_count += 1

        raw_doi = str(split_row[19]) if split_row[19] else ""
        doi_fn = _clean_doi_filename(raw_doi) if raw_doi else ""
        if doi_fn:
            local = OPEX_DOCS / "invoice delivery file" / doi_fn
            key = "po/do_invoices/" + doi_fn.replace(" ", "_")
            if _upload_file(cur, local, key, split_po, "delivery_order_invoice_file"):
                doi_count += 1

    conn.close()
    print(f"\n[OK] Attached: {po_inv_count} PO invoices, {do_count} resi, {doi_count} DO invoices")


# ---------------------------------------------------------------------------
# Step 5b — Split delivery POs
# ---------------------------------------------------------------------------
# POs 1, 2, and 8 each had two physical shipments. The per-item delivery date
# in the PO-specific recap sheets identifies which items belong to each batch.
# Items with the split_date go to the split PO; remaining (or null date) stay
# in the parent PO.


def _new_ulid() -> str:
    return str(ulid_lib.new().uuid)


SPLIT_DELIVERY_CONFIG = [
    {
        "parent_po": "PO-2025-001",
        "split_po": "PO-2025-001B",
        "sheet": "PO Recap 2 April 2025",
        "sku_col": 8,
        "date_col": 16,
        "split_date": date(2025, 5, 15),
        "tracker_no": 1,
    },
    {
        "parent_po": "PO-2025-002",
        "split_po": "PO-2025-002B",
        "sheet": "PO Recap 2 Juni 2025",
        "sku_col": 9,
        "date_col": 17,
        "split_date": date(2025, 8, 15),
        "tracker_no": 2,
    },
    {
        "parent_po": "PO-2026-001",
        "split_po": "PO-2026-001B",
        "sheet": "Recap 9 Jan 2026",
        "sku_col": 9,
        "date_col": 17,
        "split_date": date(2026, 5, 7),
        "tracker_no": 8,
    },
]


def step_split_deliveries() -> None:
    header("Step 5b: Split delivery POs")

    # Load tracker (need second row per split no)
    wb_tracker = openpyxl.load_workbook(
        MIRAKO_DATA / "Purchase Order Tracker (1).xlsx", data_only=True
    )
    ws_tracker = wb_tracker["Master Tracker Purchase Order"]
    tracker_rows = list(ws_tracker.iter_rows(values_only=True))

    # Group tracker rows by "no" (preserving order → second row is the split)
    by_no: dict[int, list] = defaultdict(list)
    for row in tracker_rows[7:]:
        no = row[0]
        if isinstance(no, (int, float)) and int(no) in TRACKER_NO_TO_PO:
            by_no[int(no)].append(row)

    # Load PO detail Excel
    wb_detail = openpyxl.load_workbook(
        MIRAKO_DATA / "Purchase Order Detail Overview (1).xlsx", data_only=True
    )

    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()

    # Disable PO number auto-generation trigger so we can set explicit PO numbers
    cur.execute("ALTER TABLE purchasing_purchaseorder DISABLE TRIGGER trg_generate_po_number")

    for cfg in SPLIT_DELIVERY_CONFIG:
        parent_po = cfg["parent_po"]
        split_po = cfg["split_po"]
        tracker_no = cfg["tracker_no"]
        split_date = cfg["split_date"]

        # Fetch parent PO metadata from DB
        cur.execute(
            """
            SELECT purchase_order_id, company_id, warehouse_id, supplier_id,
                   exchange_rate, invoice_number, invoice_date, commission_fee_pct
            FROM purchasing_purchaseorder
            WHERE purchase_order_number = %s
            """,
            (parent_po,),
        )
        parent_row = cur.fetchone()
        if not parent_row:
            print(f"  SKIP {split_po}: parent {parent_po} not found")
            continue
        (
            parent_po_id,
            company_id,
            warehouse_id,
            supplier_id,
            exchange_rate,
            invoice_number,
            invoice_date,
            commission_fee_pct,
        ) = parent_row
        # Get second tracker row for this no (the split delivery)
        rows_for_no = by_no.get(tracker_no, [])
        if len(rows_for_no) < 2:
            print(f"  SKIP {split_po}: no second tracker row for no={tracker_no}")
            continue
        split_row = rows_for_no[1]

        # Parse split RESI
        raw_do = str(split_row[4]) if split_row[4] else ""
        resi_filename = _clean_resi_filename(raw_do) if raw_do else ""
        resi_number = re.sub(r"\.pdf.*$", "", resi_filename, flags=re.IGNORECASE).strip()

        # Parse split DO invoice
        raw_doi = str(split_row[19]) if split_row[19] else ""
        doi_filename = _clean_doi_filename(raw_doi) if raw_doi else ""

        # delivery_fee for split row (col 30 = shipping_1688 in RMB)
        delivery_fee_rmb = round(float(split_row[30] or 0), 3)
        shipping_fee = int(split_row[16] or 0)
        delivery_date = _to_date(split_row[5])
        created_date = _to_date(split_row[1])

        # Read per-item delivery dates from PO sheet to find split variants
        ws = wb_detail[cfg["sheet"]]
        sheet_rows = list(ws.iter_rows(values_only=True))
        split_variant_skus: set[str] = set()
        for row in sheet_rows[2:]:
            sku = row[cfg["sku_col"]]
            item_date = row[cfg["date_col"]]
            if not sku:
                continue
            item_date_d = _to_date(item_date)
            if item_date_d == split_date:
                split_variant_skus.add(str(sku).strip())

        if not split_variant_skus:
            print(f"  SKIP {split_po}: no items with delivery date {split_date}")
            continue

        print(f"  {split_po}: {len(split_variant_skus)} variant SKUs for split date {split_date}")

        # Resolve variant IDs from DB
        cur.execute(
            """
            SELECT pv.product_variant_id, pv.sku_variant_code
            FROM master_productvariant pv
            WHERE pv.sku_variant_code = ANY(%s)
            """,
            (list(split_variant_skus),),
        )
        variant_rows = cur.fetchall()
        split_variant_ids = [r[0] for r in variant_rows]
        found_skus = {r[1] for r in variant_rows}
        missing = split_variant_skus - found_skus
        if missing:
            print(f"    WARNING: {len(missing)} SKUs not in DB: {list(missing)[:5]}")
        if not split_variant_ids:
            print(f"  SKIP {split_po}: no variant IDs resolved")
            continue

        # Count how many details will move
        cur.execute(
            """
            SELECT COUNT(*) FROM purchasing_purchaseorderdetail
            WHERE purchase_order_id = %s AND product_variant_id = ANY(%s)
            """,
            (parent_po_id, split_variant_ids),
        )
        n_moving = cur.fetchone()[0]
        print(f"    Moving {n_moving} PO detail rows to {split_po}")

        # Create the split PO record (copy metadata from parent, override delivery fields)
        split_po_id = _new_ulid()
        from datetime import (
            datetime as _dt,
            timezone as _tz,
        )

        ts = _dt.now(_tz.utc)
        created_ts = (
            _dt.combine(created_date, _dt.min.time()).replace(tzinfo=_tz.utc)
            if created_date
            else ts
        )
        cur.execute(
            """
            INSERT INTO purchasing_purchaseorder (
                purchase_order_id, company_id, purchase_order_number, status,
                invoice_number, invoice_date, delivery_order_number, delivery_date,
                warehouse_id, supplier_id, supplier_name,
                currency, exchange_rate,
                commission_fee_pct, delivery_fee, shipping_fee,
                total_ordered_qty, total_received_qty,
                total_item_amount, procure_amount, total_order_amount, total_amount,
                has_discount, cdate, udate
            ) VALUES (
                %s, %s, %s, 'COMPLETED',
                %s, %s, %s, %s,
                %s, %s, 'Ancorelala',
                'CNY', %s,
                %s, %s, %s,
                0, 0,
                0, 0, 0, 0,
                FALSE, %s, %s
            )
            """,
            (
                split_po_id,
                company_id,
                split_po,
                invoice_number,
                invoice_date,
                resi_number or None,
                delivery_date,
                warehouse_id,
                supplier_id,
                exchange_rate,
                commission_fee_pct,
                delivery_fee_rmb,
                shipping_fee,
                created_ts,
                ts,
            ),
        )

        # Move matching PO details to the split PO
        cur.execute(
            """
            UPDATE purchasing_purchaseorderdetail
            SET purchase_order_id = %s
            WHERE purchase_order_id = %s AND product_variant_id = ANY(%s)
            """,
            (split_po_id, parent_po_id, split_variant_ids),
        )

        # Update FIFO reference_number for moved variants
        cur.execute(
            """
            UPDATE inventory_productcogs
            SET reference_number = %s
            WHERE reference_number = %s AND product_variant_id = ANY(%s)
            """,
            (split_po, parent_po, split_variant_ids),
        )

        # Recalculate totals for both POs from their remaining details
        for po_id, po_num in [(parent_po_id, parent_po), (split_po_id, split_po)]:
            cur.execute(
                """
                SELECT
                    COALESCE(SUM(ordered_qty), 0),
                    COALESCE(SUM(total_price_base), 0)
                FROM purchasing_purchaseorderdetail
                WHERE purchase_order_id = %s
                """,
                (po_id,),
            )
            qty_total, item_amount = cur.fetchone()
            cur.execute(
                "UPDATE purchasing_purchaseorder SET "
                "total_ordered_qty = %s, total_received_qty = %s, "
                "total_item_amount = %s, procure_amount = %s "
                "WHERE purchase_order_id = %s",
                (qty_total, qty_total, item_amount, item_amount, po_id),
            )

        print(
            f"  ✓ Created {split_po} (RESI={resi_number or 'none'}, DO_inv={doi_filename or 'none'})"
        )

    # Re-enable trigger
    cur.execute("ALTER TABLE purchasing_purchaseorder ENABLE TRIGGER trg_generate_po_number")
    conn.close()
    print("\n[OK] Split delivery POs created")


# ---------------------------------------------------------------------------
# Step 9 — Reconciliation report
# ---------------------------------------------------------------------------
def step_reconciliation() -> None:
    header("Step 9: Reconciliation")
    conn = psycopg.connect(**DB_PARAMS, autocommit=True)
    cur = conn.cursor()

    checks = [
        ("Products", "SELECT COUNT(*) FROM master_product"),
        ("Variants", "SELECT COUNT(*) FROM master_productvariant"),
        ("Purchase Orders", "SELECT COUNT(*) FROM purchasing_purchaseorder"),
        ("PO Details", "SELECT COUNT(*) FROM purchasing_purchaseorderdetail"),
        ("FIFO records", "SELECT COUNT(*) FROM inventory_productcogs"),
        ("Sales Orders", "SELECT COUNT(*) FROM sales_salesorder"),
        ("Sales Items", "SELECT COUNT(*) FROM sales_salesorderitem"),
        ("Cash Transactions", "SELECT COUNT(*) FROM finance_cashtransaction"),
        (
            "POs with invoice_number",
            "SELECT COUNT(*) FROM purchasing_purchaseorder WHERE invoice_number IS NOT NULL",
        ),
        (
            "POs with delivery_order_number",
            "SELECT COUNT(*) FROM purchasing_purchaseorder WHERE delivery_order_number IS NOT NULL",
        ),
        (
            "POs with PO invoice file",
            "SELECT COUNT(*) FROM purchasing_purchaseorder WHERE purchase_order_invoice_file IS NOT NULL",
        ),
        (
            "POs with resi file",
            "SELECT COUNT(*) FROM purchasing_purchaseorder WHERE delivery_order_file IS NOT NULL",
        ),
        (
            "POs with DO invoice file",
            "SELECT COUNT(*) FROM purchasing_purchaseorder WHERE delivery_order_invoice_file IS NOT NULL",
        ),
        (
            "Product-supplier links",
            "SELECT COUNT(*) FROM inventory_productsupplier WHERE supplier_link IS NOT NULL AND supplier_link != ''",
        ),
        (
            "Negative FIFO qty",
            "SELECT COUNT(*) FROM inventory_productcogs WHERE remaining_qty < 0",
        ),
    ]

    all_ok = True
    for label, sql in checks:
        cur.execute(sql)
        count = cur.fetchone()[0]
        status = "[WARN]" if (label == "Negative FIFO qty" and count > 0) else "[OK]  "
        if "WARN" in status:
            all_ok = False
        print(f"  {status} {label}: {count}")

    conn.close()
    print(f"\n{'All checks passed' if all_ok else 'Some warnings — review above'}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Mirako full data migration")
    parser.add_argument("--skip-files", action="store_true", help="Skip R2 file uploads")
    args = parser.parse_args()

    print("\nMirako Data Migration")
    print(f"MIRAKO_DATA: {MIRAKO_DATA}")
    if not MIRAKO_DATA.exists():
        print(f"ERROR: mirako_data folder not found at {MIRAKO_DATA}")
        sys.exit(1)

    step_reset()
    step_seed_user()
    step_master_data()
    step_product_links()
    step_purchase_orders()
    step_enrich_po_tracker()
    step_split_deliveries()
    step_sales_orders()
    step_cash_transactions()
    step_attach_files(args.skip_files)
    step_reconciliation()

    print("\n" + "=" * 60)
    print("  Migration complete.")
    print("=" * 60)


if __name__ == "__main__":
    main()
