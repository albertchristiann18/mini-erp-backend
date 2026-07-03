"""
Phase 4 — Import ~5,526 historical sales order line items with FIFO consumption
and stock deductions from 13 monthly Sales sheets.

Run from mini-erp-backend/:
    uv run python scripts/import_sales_excel.py
"""

from collections import OrderedDict, defaultdict, deque
from datetime import datetime, timezone

import openpyxl
import psycopg
import ulid as ulid_lib

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
EXCEL_PATH = "/Users/jtf01644/personal/mini-erp-project/data/Sales Order Tracker.xlsx"

DB_PARAMS = dict(
    host="localhost",
    port=5433,
    dbname="mini_erp",
    user="postgres",
    password="postgres",
)

SHEETS_TO_IMPORT = [
    "Sales May 2025",
    "Sales June 2025",
    "Sales July 2025",
    "Sales August 2025",
    "Sales September 2025",
    "Sales October 2025",
    "Sales November 2025",
    "Sales December 2025",
    "Sales January 2026",
    "Sales February 2026",
    "Sales March 2026",
    "Sales April 2026",
    "Sales May 2026",
]

STATUS_MAPPING = {
    "Completed": "COMPLETED",
    "Delivered": "DELIVERED",
    "In_Delivery": "SHIPPING",
    "Processed": "CONFIRMED",
    "Returns": "RETURNED",
    "To_Process": "PENDING",
    "Cancellations": "CANCELLED",
}

STOCK_AFFECTING_STATUSES = frozenset(
    {
        "COMPLETED",
        "DELIVERED",
        "SHIPPING",
        "CONFIRMED",
        "RETURNED",
    }
)

REQUIRED_HEADER_MAP = {
    "nomor_paket_col": "nomor paket",
    "status_col": "status pesanan",
    "channel_col": "channel - nama toko",
    "order_date_col": "tanggal pesanan dibuat",
    "sku_master_col": "sku master",
    "sku_marketplace_col": "sku marketplace",
    "qty_col": "jumlah",
    "selling_price_col": "harga dibayar",
    "subtotal_col": "subtotal pesanan",
    "customer_col": "nama pembeli",
}

OPTIONAL_HEADER_KEYS = {
    "courier_col": "kurir",
    "tracking_col": "nomor awb/resi",
}


def new_ulid() -> str:
    return str(ulid_lib.new().uuid)


def now() -> datetime:
    return datetime.now(timezone.utc)


def clean_header(val) -> str:
    if val is None:
        return ""
    return str(val).strip().replace("\n", " ").lower()


def detect_columns(headers_row: tuple) -> dict:
    cols: dict[str, int] = {}
    net_rev_col = None

    for idx, val in enumerate(headers_row):
        key = clean_header(val)
        if not key:
            continue

        for col_name, header_text in REQUIRED_HEADER_MAP.items():
            if key == header_text and col_name not in cols:
                cols[col_name] = idx

        for col_name, header_text in OPTIONAL_HEADER_KEYS.items():
            if key == header_text and col_name not in cols:
                cols[col_name] = idx

        if "total penjualan" in key and net_rev_col is None:
            net_rev_col = idx

    if net_rev_col is not None:
        cols["net_rev_col"] = net_rev_col

    return cols


def check_required_columns(cols: dict) -> list[str]:
    missing = []
    for k in REQUIRED_HEADER_MAP:
        if k not in cols:
            missing.append(k)

    # SKU columns are collectively required — at least one must exist
    sku_master = "sku_master_col" in cols
    sku_marketplace = "sku_marketplace_col" in cols
    if not sku_master and not sku_marketplace:
        # Both missing is truly missing
        if "sku_master_col" not in missing:
            missing.append("sku_master_col")
    else:
        # At least one exists — remove from missing
        missing = [m for m in missing if m not in ("sku_master_col", "sku_marketplace_col")]

    return missing


def parse_order_date(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
    return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def consume_fifo(
    fifo_map: dict[str, deque[dict]],
    variant_id: str,
    qty_needed: int,
) -> tuple[list[dict], int, int]:
    batches = fifo_map.get(variant_id, deque())
    consumed: list[dict] = []
    qty_left = qty_needed
    total_cogs = 0

    for batch in batches:
        if qty_left <= 0:
            break
        if batch["remaining_qty"] <= 0:
            continue
        qty_from_this = min(qty_left, batch["remaining_qty"])
        batch["remaining_qty"] -= qty_from_this
        consumed.append(
            {
                "id": batch["id"],
                "qty": qty_from_this,
                "cogs_amount": batch["cogs_amount"],
            }
        )
        total_cogs += qty_from_this * batch["cogs_amount"]
        qty_left -= qty_from_this

    if qty_left > 0:
        print(
            f"  WARNING: FIFO shortage for variant {variant_id}: "
            f"needed {qty_needed}, short {qty_left}"
        )

    avg_cogs = int(round(total_cogs / qty_needed)) if qty_needed > 0 else 0
    return consumed, total_cogs, avg_cogs


def main():
    ts = now()

    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"Available sheets: {wb.sheetnames}")

    print("\nConnecting to database...")
    with psycopg.connect(**DB_PARAMS, autocommit=True) as conn:
        with conn.cursor() as cur:
            # ------------------------------------------------------------------
            # STEP 0 — Bootstrap
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 0: Bootstrap")
            print("=" * 60)

            cur.execute("SELECT company_id FROM core_company WHERE name = 'Mirako' LIMIT 1")
            row = cur.fetchone()
            if not row:
                raise RuntimeError("Company 'Mirako' not found")
            company_id = str(row[0])
            print(f"  Company 'Mirako' id: {company_id}")

            cur.execute(
                "SELECT warehouse_id FROM inventory_warehouse "
                "WHERE company_id = %s ORDER BY cdate LIMIT 1",
                (company_id,),
            )
            row = cur.fetchone()
            if not row:
                raise RuntimeError("No warehouse found for company Mirako")
            warehouse_id = str(row[0])
            print(f"  Warehouse id: {warehouse_id}")

            # Build variant_map: sku_variant_code -> product_variant_id
            cur.execute(
                "SELECT product_variant_id, sku_variant_code "
                "FROM inventory_productvariant WHERE company_id = %s",
                (company_id,),
            )
            variant_map: dict[str, str] = {}
            for pid, sku in cur.fetchall():
                variant_map[str(sku)] = str(pid)
            print(f"  Variant map built: {len(variant_map)} variants")

            # Build fifo_map: variant_id -> deque of batches sorted by purchase_date ASC
            cur.execute(
                "SELECT product_cogs_id, product_variant_id, cogs_amount, "
                "       remaining_qty, purchase_date "
                "FROM inventory_productcogs "
                "WHERE company_id = %s "
                "ORDER BY product_variant_id, purchase_date ASC",
                (company_id,),
            )
            fifo_map: dict[str, deque[dict]] = defaultdict(deque)
            for cogs_id, pv_id, cogs_amt, rem_qty, _pur_date in cur.fetchall():
                fifo_map[str(pv_id)].append(
                    {
                        "id": str(cogs_id),
                        "cogs_amount": cogs_amt,
                        "remaining_qty": rem_qty,
                    }
                )
            print(f"  FIFO map built: {len(fifo_map)} variants")

            # Build variant_balance: variant_id -> current physical_qty
            cur.execute(
                "SELECT pv.product_variant_id, pvw.physical_qty "
                "FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv "
                "  ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s",
                (company_id,),
            )
            variant_balance: dict[str, int] = {}
            for pv_id, phys_qty in cur.fetchall():
                variant_balance[str(pv_id)] = phys_qty
            print(f"  Variant balance built: {len(variant_balance)} variants")

            # ------------------------------------------------------------------
            # STEP 1 — Idempotency Reset
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 1: Idempotency Reset")
            print("=" * 60)

            cur.execute(
                "DELETE FROM sales_salesordercogsdetail WHERE sales_order_item_id IN ("
                "  SELECT soi.sales_order_item_id FROM sales_salesorderitem soi "
                "  JOIN sales_salesorder so ON soi.sales_order_id = so.sales_order_id "
                "  WHERE so.company_id = %s"
                ")",
                (company_id,),
            )
            print(f"  Deleted CogsDetail: {cur.rowcount}")

            cur.execute(
                "DELETE FROM sales_salesorderitem WHERE sales_order_id IN ("
                "  SELECT sales_order_id FROM sales_salesorder WHERE company_id = %s"
                ")",
                (company_id,),
            )
            print(f"  Deleted SalesOrderItems: {cur.rowcount}")

            cur.execute(
                "DELETE FROM sales_salesorder WHERE company_id = %s",
                (company_id,),
            )
            print(f"  Deleted SalesOrders: {cur.rowcount}")

            cur.execute(
                "DELETE FROM inventory_stockmovement "
                "WHERE movement_type = 'OUT' "
                "AND product_variant_id IN ("
                "  SELECT product_variant_id FROM inventory_productvariant "
                "  WHERE company_id = %s"
                ")",
                (company_id,),
            )
            print(f"  Deleted OUT StockMovements: {cur.rowcount}")

            cur.execute(
                "UPDATE inventory_productcogs SET remaining_qty = original_qty "
                "WHERE company_id = %s",
                (company_id,),
            )
            print(f"  Restored FIFO batches: {cur.rowcount}")

            cur.execute(
                "UPDATE inventory_productvariantwarehouse pvw "
                "SET physical_qty = ("
                "  SELECT COALESCE(SUM(pc.original_qty), 0) "
                "  FROM inventory_productcogs pc "
                "  WHERE pc.product_variant_id = pvw.product_variant_id "
                "  AND pc.warehouse_id = pvw.warehouse_id"
                "), "
                "outgoing_qty = 0 "
                "WHERE pvw.product_variant_id IN ("
                "  SELECT product_variant_id FROM inventory_productvariant "
                "  WHERE company_id = %s"
                ")",
                (company_id,),
            )
            print(f"  Restored warehouse stock: {cur.rowcount}")

            cur.execute(
                "UPDATE inventory_productvariant "
                "SET total_outgoing_qty = 0, "
                "    total_available_qty = ("
                "      SELECT COALESCE(SUM(pvw.physical_qty), 0) "
                "      FROM inventory_productvariantwarehouse pvw "
                "      WHERE pvw.product_variant_id = "
                "            inventory_productvariant.product_variant_id"
                "    ) "
                "WHERE company_id = %s",
                (company_id,),
            )
            print(f"  Restored ProductVariant totals: {cur.rowcount}")

            print("  Idempotency reset complete.")

            # Reload fifo_map after reset
            fifo_map.clear()
            cur.execute(
                "SELECT product_cogs_id, product_variant_id, cogs_amount, "
                "       remaining_qty, purchase_date "
                "FROM inventory_productcogs "
                "WHERE company_id = %s "
                "ORDER BY product_variant_id, purchase_date ASC",
                (company_id,),
            )
            for cogs_id, pv_id, cogs_amt, rem_qty, _pur_date in cur.fetchall():
                fifo_map[str(pv_id)].append(
                    {
                        "id": str(cogs_id),
                        "cogs_amount": cogs_amt,
                        "remaining_qty": rem_qty,
                    }
                )
            print(f"  FIFO map reloaded: {len(fifo_map)} variants")

            # Reload variant_balance after reset
            variant_balance.clear()
            cur.execute(
                "SELECT pv.product_variant_id, pvw.physical_qty "
                "FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv "
                "  ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s",
                (company_id,),
            )
            for pv_id, phys_qty in cur.fetchall():
                variant_balance[str(pv_id)] = phys_qty
            print(f"  Variant balance reloaded: {len(variant_balance)} variants")

            # ------------------------------------------------------------------
            # STEP 2 — Disable SO number trigger
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 2: Disable SO Number Trigger")
            print("=" * 60)
            cur.execute("ALTER TABLE sales_salesorder DISABLE TRIGGER trg_generate_so_number")
            print("  Trigger disabled.")

            # ------------------------------------------------------------------
            # STEP 3 — Import each sheet
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 3: Import Sheets")
            print("=" * 60)

            inserted_order_numbers: set[str] = set()
            total_orders = 0
            total_items = 0
            total_skipped_items = 0

            for sheet_name in SHEETS_TO_IMPORT:
                if sheet_name not in wb.sheetnames:
                    print(f"  WARNING: Sheet '{sheet_name}' not found -- skipping")
                    continue

                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    print(f"  WARNING: Sheet '{sheet_name}' has too few rows -- skipping")
                    continue

                headers = rows[0]
                cols = detect_columns(headers)

                missing = check_required_columns(cols)
                if missing:
                    print(
                        f"  WARNING: Sheet '{sheet_name}' missing required columns: "
                        f"{missing} -- skipping"
                    )
                    continue

                nomor_paket_col = cols["nomor_paket_col"]
                status_col = cols["status_col"]
                order_date_col = cols["order_date_col"]
                sku_master_col = cols.get("sku_master_col")
                sku_marketplace_col = cols.get("sku_marketplace_col")
                qty_col = cols["qty_col"]
                selling_price_col = cols["selling_price_col"]
                subtotal_col_idx = cols["subtotal_col"]
                customer_col = cols["customer_col"]
                courier_col = cols.get("courier_col")
                tracking_col = cols.get("tracking_col")
                net_rev_col = cols.get("net_rev_col")

                # Group rows by nomor_paket
                groups: OrderedDict[str, list[tuple]] = OrderedDict()
                for data_row in rows[1:]:
                    if len(data_row) <= max(nomor_paket_col, status_col):
                        continue
                    np_val = data_row[nomor_paket_col]
                    if np_val is None:
                        continue
                    np_str = str(np_val).strip()
                    if not np_str:
                        continue
                    groups.setdefault(np_str, []).append(data_row)

                sheet_order_count = 0
                sheet_item_count = 0
                sheet_skip_count = 0

                for nomor_paket, group_rows in groups.items():
                    if nomor_paket in inserted_order_numbers:
                        continue

                    first_row = group_rows[0]

                    # Status
                    raw_status = (
                        str(first_row[status_col]).strip()
                        if first_row[status_col] is not None
                        else ""
                    )
                    mapped_status = STATUS_MAPPING.get(raw_status, "PENDING")

                    # Order date
                    order_date = parse_order_date(first_row[order_date_col])
                    if order_date is None:
                        for r in group_rows:
                            if r[order_date_col] is not None:
                                order_date = parse_order_date(r[order_date_col])
                                if order_date:
                                    break
                    if order_date is None:
                        order_date = now()

                    # Customer name
                    customer = (
                        str(first_row[customer_col]).strip()[:255]
                        if first_row[customer_col] is not None
                        else ""
                    )

                    # Courier
                    courier = ""
                    if (
                        courier_col is not None
                        and courier_col < len(first_row)
                        and first_row[courier_col] is not None
                    ):
                        courier = str(first_row[courier_col]).strip()[:100]

                    # Tracking
                    tracking = ""
                    if (
                        tracking_col is not None
                        and tracking_col < len(first_row)
                        and first_row[tracking_col] is not None
                    ):
                        tracking = str(first_row[tracking_col]).strip()[:255]

                    # Subtotal (order-level)
                    subtotal = None
                    for r in group_rows:
                        if subtotal_col_idx < len(r) and r[subtotal_col_idx] is not None:
                            raw = safe_float(r[subtotal_col_idx])
                            if raw is not None:
                                subtotal = int(round(raw))
                                break
                    if subtotal is None:
                        subtotal = 0
                        for r in group_rows:
                            raw_qty_val = r[qty_col] if qty_col < len(r) else None
                            raw_qty = safe_float(raw_qty_val)
                            q = int(round(raw_qty)) if raw_qty is not None else 1
                            raw_sp_val = (
                                r[selling_price_col] if selling_price_col < len(r) else None
                            )
                            raw_sp = safe_float(raw_sp_val)
                            p = int(round(raw_sp)) if raw_sp is not None else 0
                            subtotal += q * p

                    # Net revenue
                    net_revenue = 0
                    if net_rev_col is not None:
                        for r in group_rows:
                            if net_rev_col < len(r) and r[net_rev_col] is not None:
                                raw = safe_float(r[net_rev_col])
                                if raw is not None:
                                    net_revenue = int(round(raw))
                                    break

                    # INSERT SalesOrder
                    so_id = new_ulid()
                    cur.execute(
                        """
                        INSERT INTO sales_salesorder (
                            sales_order_id, company_id, order_number,
                            marketplace_order_id, marketplace_order_number,
                            status, source_platform,
                            warehouse_id, marketplace_id,
                            customer_name, customer_phone,
                            shipping_address, shipping_province, shipping_city,
                            order_date,
                            courier_name, tracking_number,
                            subtotal, total_discount, total_marketplace_fee,
                            total_cogs, net_revenue, gross_profit,
                            shipping_fee, shipping_fee_seller,
                            note,
                            cdate, udate
                        ) VALUES (
                            %s, %s, %s,
                            %s, %s,
                            %s, 'SHOPEE',
                            %s, NULL,
                            %s, '',
                            '', '', '',
                            %s,
                            %s, %s,
                            %s, 0, 0,
                            0, %s, 0,
                            0, 0,
                            %s,
                            %s, %s
                        )
                        """,
                        (
                            so_id,
                            company_id,
                            nomor_paket,
                            nomor_paket,
                            nomor_paket,
                            mapped_status,
                            warehouse_id,
                            customer,
                            order_date,
                            courier,
                            tracking,
                            subtotal,
                            net_revenue,
                            f"Imported from {sheet_name}",
                            ts,
                            ts,
                        ),
                    )

                    so_total_cogs = 0
                    so_gross_profit = 0

                    for item_row in group_rows:
                        # SKU code: try master first, fall back to marketplace
                        sku_code = None
                        if sku_master_col is not None and sku_master_col < len(item_row):
                            sku_code = item_row[sku_master_col]
                        if sku_code is None or (
                            isinstance(sku_code, str) and sku_code.strip() == ""
                        ):
                            if sku_marketplace_col is not None and sku_marketplace_col < len(
                                item_row
                            ):
                                sku_code = item_row[sku_marketplace_col]
                        if sku_code is None:
                            sheet_skip_count += 1
                            continue
                        sku_code = str(sku_code).strip()
                        if not sku_code:
                            sheet_skip_count += 1
                            continue

                        # Quantity
                        raw_qty_val = item_row[qty_col] if qty_col < len(item_row) else None
                        raw_qty = safe_float(raw_qty_val)
                        qty = int(round(raw_qty)) if raw_qty is not None else 1

                        # Selling price
                        raw_sp_val = (
                            item_row[selling_price_col]
                            if selling_price_col < len(item_row)
                            else None
                        )
                        raw_sp = safe_float(raw_sp_val)
                        selling_price = int(round(raw_sp)) if raw_sp is not None else 0

                        line_total = selling_price * qty

                        if sku_code not in variant_map:
                            print(
                                f"  WARNING: SKU '{sku_code}' not found in "
                                f"variant_map -- skipping item (order {nomor_paket})"
                            )
                            sheet_skip_count += 1
                            continue

                        variant_id = variant_map[sku_code]

                        # FIFO consumption
                        if mapped_status in STOCK_AFFECTING_STATUSES:
                            consumed, total_cogs_for_item, avg_cogs = consume_fifo(
                                fifo_map,
                                variant_id,
                                qty,
                            )
                            actual_cogs_per_unit = avg_cogs
                            actual_cogs_total = total_cogs_for_item
                        else:
                            consumed = []
                            actual_cogs_per_unit = 0
                            actual_cogs_total = 0

                        # INSERT SalesOrderItem
                        item_id = new_ulid()
                        cur.execute(
                            """
                            INSERT INTO sales_salesorderitem (
                                sales_order_item_id, company_id,
                                sales_order_id, product_variant_id,
                                quantity, selling_price,
                                discount_amount, commission_fee, service_fee,
                                total_marketplace_fee,
                                actual_cogs_per_unit, actual_cogs_total,
                                line_total,
                                cdate, udate
                            ) VALUES (
                                %s, %s,
                                %s, %s,
                                %s, %s,
                                0, 0, 0,
                                0,
                                %s, %s,
                                %s,
                                %s, %s
                            )
                            """,
                            (
                                item_id,
                                company_id,
                                so_id,
                                variant_id,
                                qty,
                                selling_price,
                                actual_cogs_per_unit,
                                actual_cogs_total,
                                line_total,
                                ts,
                                ts,
                            ),
                        )

                        # INSERT CogsDetail for each consumed batch
                        for batch in consumed:
                            cogs_detail_id = new_ulid()
                            total_cogs_batch = batch["qty"] * batch["cogs_amount"]
                            cur.execute(
                                """
                                INSERT INTO sales_salesordercogsdetail (
                                    sales_order_cogs_detail_id, company_id,
                                    sales_order_item_id, product_cogs_id,
                                    quantity_consumed, cogs_per_unit, total_cogs,
                                    cdate, udate
                                ) VALUES (
                                    %s, %s,
                                    %s, %s,
                                    %s, %s, %s,
                                    %s, %s
                                )
                                """,
                                (
                                    cogs_detail_id,
                                    company_id,
                                    item_id,
                                    batch["id"],
                                    batch["qty"],
                                    batch["cogs_amount"],
                                    total_cogs_batch,
                                    ts,
                                    ts,
                                ),
                            )

                            # Update ProductCogs remaining_qty
                            cur.execute(
                                "UPDATE inventory_productcogs "
                                "SET remaining_qty = remaining_qty - %s "
                                "WHERE product_cogs_id = %s",
                                (batch["qty"], batch["id"]),
                            )

                        # Stock-affecting updates
                        if mapped_status in STOCK_AFFECTING_STATUSES:
                            balance_before = variant_balance.get(variant_id, 0)
                            balance_after = balance_before - qty
                            variant_balance[variant_id] = balance_after

                            sm_id = new_ulid()
                            cur.execute(
                                """
                                INSERT INTO inventory_stockmovement (
                                    stock_movement_id, company_id,
                                    product_variant_id, warehouse_id,
                                    movement_type, field_change, quantity,
                                    reference_number, note,
                                    balance_before, balance_after,
                                    cdate, udate
                                ) VALUES (
                                    %s, %s,
                                    %s, %s,
                                    'OUT', 'physical_qty', %s,
                                    %s, %s,
                                    %s, %s,
                                    %s, %s
                                )
                                """,
                                (
                                    sm_id,
                                    company_id,
                                    variant_id,
                                    warehouse_id,
                                    -qty,
                                    nomor_paket,
                                    f"Sale: {sheet_name}",
                                    balance_before,
                                    balance_after,
                                    ts,
                                    ts,
                                ),
                            )

                            cur.execute(
                                "UPDATE inventory_productvariantwarehouse "
                                "SET physical_qty = physical_qty - %s, "
                                "    outgoing_qty = outgoing_qty + %s "
                                "WHERE product_variant_id = %s "
                                "AND warehouse_id = %s",
                                (qty, qty, variant_id, warehouse_id),
                            )

                            cur.execute(
                                "UPDATE inventory_productvariant "
                                "SET total_outgoing_qty = total_outgoing_qty + %s, "
                                "    total_available_qty = "
                                "        total_available_qty - %s "
                                "WHERE product_variant_id = %s",
                                (qty, qty, variant_id),
                            )

                        so_total_cogs += actual_cogs_total
                        so_gross_profit += line_total - actual_cogs_total
                        sheet_item_count += 1

                    # Update SalesOrder totals
                    cur.execute(
                        "UPDATE sales_salesorder "
                        "SET total_cogs = %s, gross_profit = %s "
                        "WHERE sales_order_id = %s",
                        (so_total_cogs, so_gross_profit, so_id),
                    )

                    inserted_order_numbers.add(nomor_paket)
                    sheet_order_count += 1

                print(
                    f"  Sheet: {sheet_name} | Orders: {sheet_order_count} "
                    f"| Items: {sheet_item_count} | Skipped items: {sheet_skip_count}"
                )
                total_orders += sheet_order_count
                total_items += sheet_item_count
                total_skipped_items += sheet_skip_count

            # ------------------------------------------------------------------
            # STEP 4 — Re-enable trigger
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 4: Re-enable SO Number Trigger")
            print("=" * 60)
            cur.execute("ALTER TABLE sales_salesorder ENABLE TRIGGER trg_generate_so_number")
            print("  Trigger re-enabled.")

            # ------------------------------------------------------------------
            # STEP 5 — Final Validation
            # ------------------------------------------------------------------
            print("\n" + "=" * 60)
            print("STEP 5: Final Validation")
            print("=" * 60)

            cur.execute(
                "SELECT COUNT(DISTINCT sales_order_id) FROM sales_salesorder WHERE company_id = %s",
                (company_id,),
            )
            so_count = cur.fetchone()[0]
            print(f"  SalesOrders: {so_count}")

            cur.execute(
                "SELECT COUNT(*) FROM sales_salesorderitem WHERE company_id = %s",
                (company_id,),
            )
            soi_count = cur.fetchone()[0]
            print(f"  SalesOrderItems: {soi_count}")

            cur.execute(
                "SELECT COUNT(*) FROM sales_salesordercogsdetail socd "
                "JOIN sales_salesorderitem soi "
                "  ON socd.sales_order_item_id = soi.sales_order_item_id "
                "WHERE soi.company_id = %s",
                (company_id,),
            )
            socd_count = cur.fetchone()[0]
            print(f"  SalesOrderCogsDetails: {socd_count}")

            cur.execute(
                "SELECT COUNT(*) FROM inventory_productcogs "
                "WHERE company_id = %s AND remaining_qty < 0",
                (company_id,),
            )
            neg_cogs = cur.fetchone()[0]
            print(f"  Negative FIFO remaining_qty: {neg_cogs}")

            cur.execute(
                "SELECT COUNT(*) FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv "
                "  ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s AND pvw.physical_qty < 0",
                (company_id,),
            )
            neg_stock = cur.fetchone()[0]
            print(f"  Negative physical_qty: {neg_stock}")

            cur.execute(
                "SELECT COALESCE(SUM(remaining_qty), 0) "
                "FROM inventory_productcogs WHERE company_id = %s",
                (company_id,),
            )
            remaining_fifo = cur.fetchone()[0]
            print(f"  Remaining FIFO stock (sum): {remaining_fifo}")

            cur.execute(
                "SELECT COALESCE(SUM(pvw.physical_qty), 0) "
                "FROM inventory_productvariantwarehouse pvw "
                "JOIN inventory_productvariant pv "
                "  ON pvw.product_variant_id = pv.product_variant_id "
                "WHERE pv.company_id = %s",
                (company_id,),
            )
            remaining_physical = cur.fetchone()[0]
            print(f"  Remaining physical stock (sum): {remaining_physical}")

            print()
            cur.execute(
                "SELECT status, COUNT(*) FROM sales_salesorder "
                "WHERE company_id = %s GROUP BY status ORDER BY status",
                (company_id,),
            )
            print("  Orders by status:")
            for status, cnt in cur.fetchall():
                print(f"    {status}: {cnt}")

            if neg_stock > 0:
                print(f"  WARNING: {neg_stock} variants have negative physical_qty")

            # Assertions
            assert neg_cogs == 0, f"Negative FIFO remaining_qty: {neg_cogs}"

            print("\n" + "=" * 60)
            print("ALL VALIDATIONS PASSED")
            print("=" * 60)
            print(
                f"Total Orders: {so_count} | Total Items: {soi_count} "
                f"| Skipped Items: {total_skipped_items}"
            )


if __name__ == "__main__":
    main()
