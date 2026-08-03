import io
import logging
from decimal import Decimal, InvalidOperation

import openpyxl

from apps.catalog.models import Category, ProductVariant
from apps.purchasing.models import ColorAbbreviation
from core.models import Company

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"unit_price"}

COLOR_LIKE_KEYS = {"color", "warna", "colour"}


def generate_variant_suffix(
    dim1_key: str,
    dim1_value: str,
    dim2_key: str,
    dim2_value: str,
    color_map: dict[str, str],
) -> str:
    """
    Build the suffix portion of a sku_variant_code from dim values.
    Suffix order: dim2 (Size) first, dim1 (Color) second, matching SKU design.
    Color-like keys use the ColorAbbreviation map; other keys use value as-is.
    """

    def abbrev(key: str, value: str) -> str:
        if key.lower() in COLOR_LIKE_KEYS:
            return color_map.get(value.lower(), value).upper()
        return value.upper()

    parts = []
    if dim2_value:
        parts.append(abbrev(dim2_key, dim2_value))
    if dim1_value:
        parts.append(abbrev(dim1_key, dim1_value))
    return "-".join(parts)


class SourcingService:
    def build_template_workbook(self, company: Company) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()

        ws_items = wb.active
        ws_items.title = "Items"
        ws_items.append(
            [
                "variant_code",
                "product_name",
                "dim1_key",
                "dim1_value",
                "dim2_key",
                "dim2_value",
                "category_code",
                "unit_price",
                "discounted_price",
                "order_qty",
                "supplier_link",
                "image_url",
                "notes",
            ]
        )

        ws_cats = wb.create_sheet("Categories")
        ws_cats.append(["category_code", "category_name"])
        for cat in Category.objects.filter(company=company, is_active=True).order_by("name"):
            ws_cats.append([cat.category_code, cat.name])

        ws_guide = wb.create_sheet("Guide")

        ws_guide.column_dimensions["A"].width = 20
        ws_guide.column_dimensions["B"].width = 60

        guide_rows = [
            ["PANDUAN IMPORT SOURCING POOL", ""],
            ["", ""],
            ["KOLOM", "KETERANGAN"],
            [
                "variant_code",
                "Opsional. SKU varian yang sudah ada di sistem. Jika diisi, baris ini ditautkan ke produk yang sudah ada.",
            ],
            [
                "product_name",
                "Nama produk. Wajib jika supplier_link kosong. Semua baris dengan product_name yang sama dikelompokkan sebagai satu produk.",
            ],
            [
                "dim1_key",
                "Nama dimensi pertama, mis. 'Warna' (opsional — kosongkan untuk produk 1 varian).",
            ],
            [
                "dim1_value",
                "Nilai dimensi pertama, mis. 'Putih'.",
            ],
            [
                "dim2_key",
                "Nama dimensi kedua, mis. 'Ukuran' (opsional).",
            ],
            [
                "dim2_value",
                "Nilai dimensi kedua, mis. 'S'.",
            ],
            [
                "category_code",
                "Kode kategori dari sheet Categories. Boleh kosong — kategori baru akan dibuat otomatis saat import.",
            ],
            ["unit_price", "Harga beli per unit (angka saja, mis. 50000)."],
            ["discounted_price", "Harga diskon. Harus lebih kecil dari unit_price. Boleh kosong."],
            ["order_qty", "Jumlah yang disarankan untuk dipesan. Boleh kosong."],
            ["supplier_link", "URL halaman produk supplier. Wajib jika product_name kosong."],
            ["image_url", "URL gambar produk. Gambar akan diunduh otomatis. Boleh kosong."],
            ["notes", "Catatan tambahan. Boleh kosong."],
            ["", ""],
            ["CATATAN: variant_name TIDAK LAGI DIIMPUT MANUAL", ""],
            ["variant_name sekarang dihasilkan otomatis dari nilai dimensi:", ""],
            [
                "  - Jika dim1 + dim2 diisi: variant_name = '{dim1_value}-{dim2_value}' (mis. Putih-S)",
                "",
            ],
            ["  - Jika hanya dim1 diisi: variant_name = dim1_value (mis. Putih)", ""],
            ["  - Jika hanya dim2 diisi: variant_name = dim2_value", ""],
            ["", ""],
            ["CONTOH: PRODUK DENGAN 2 DIMENSI (Warna + Ukuran)", ""],
            ["", ""],
            ["Untuk produk 'Kaos Polo' dengan Warna (Putih, Merah) dan Ukuran (S, M, L):", ""],
            ["Buat 6 baris dengan product_name yang sama dan kombinasi dim yang berbeda:", ""],
            ["", ""],
            [
                "product_name | dim1_key | dim1_value | dim2_key | dim2_value | unit_price | order_qty"
            ],
            ["Kaos Polo   | Warna    | Putih      | Ukuran   | S          | 50000      | 5"],
            ["Kaos Polo   | Warna    | Putih      | Ukuran   | M          | 50000      | 10"],
            ["Kaos Polo   | Warna    | Putih      | Ukuran   | L          | 50000      | 8"],
            ["Kaos Polo   | Warna    | Merah      | Ukuran   | S          | 50000      | 5"],
            ["Kaos Polo   | Warna    | Merah      | Ukuran   | M          | 50000      | 10"],
            ["Kaos Polo   | Warna    | Merah      | Ukuran   | L          | 50000      | 8"],
            ["", ""],
            ["CONTOH: PRODUK DENGAN 1 DIMENSI (hanya Ukuran)", ""],
            ["", ""],
            [
                "product_name | dim1_key | dim1_value | dim2_key | dim2_value | unit_price | order_qty"
            ],
            ["Celana Cargo | Ukuran   | S          |          |            | 80000      | 5"],
            ["Celana Cargo | Ukuran   | M          |          |            | 80000      | 10"],
            ["Celana Cargo | Ukuran   | L          |          |            | 80000      | 8"],
            ["", ""],
            ["CONTOH: TANPA DIMENSI (produk 1 varian)", ""],
            ["", ""],
            ["Cukup isi unit_price + supplier_link, biarkan kolom dim kosong:", ""],
            ["product_name | unit_price | supplier_link"],
            ["Produk Tunggal | 50000 | https://...", ""],
        ]
        for row_data in guide_rows:
            ws_guide.append(row_data)

        return wb

    def parse_excel_preview(self, file_bytes: bytes, company: Company) -> dict[str, list[dict]]:
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(file_bytes), read_only=True, data_only=True
            )
        except Exception:
            return {
                "valid": [],
                "errors": [{"row": 0, "message": "File is not a valid Excel workbook (.xlsx)."}],
            }

        if "Items" not in wb.sheetnames:
            return {
                "valid": [],
                "errors": [{"row": 0, "message": "Sheet named 'Items' not found."}],
            }

        ws = wb["Items"]
        rows = list(ws.iter_rows(values_only=True))
        MAX_ROWS = 5000
        if len(rows) > MAX_ROWS + 1:
            return {
                "valid": [],
                "errors": [
                    {
                        "row": 0,
                        "message": f"File exceeds the {MAX_ROWS}-row limit. Split into smaller files.",
                    }
                ],
            }
        if not rows:
            return {"valid": [], "errors": [{"row": 0, "message": "Items sheet is empty."}]}

        header_row = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
        missing_required = REQUIRED_COLUMNS - set(header_row)
        if missing_required:
            return {
                "valid": [],
                "errors": [
                    {
                        "row": 1,
                        "message": f"Missing required columns: {', '.join(sorted(missing_required))}",
                    }
                ],
            }

        col_index: dict[str, int] = {col: header_row.index(col) for col in header_row if col}

        category_map: dict[str, dict[str, str]] = {
            cat.category_code: {"id": str(cat.id), "name": cat.name}
            for cat in Category.objects.filter(company=company, is_active=True)
        }

        _vc_idx = col_index.get("variant_code")
        all_variant_codes: set[str] = set()
        for _row in rows[1:]:
            if _vc_idx is not None and _vc_idx < len(_row) and _row[_vc_idx]:
                _vc_raw = str(_row[_vc_idx]).strip()
                if _vc_raw:
                    all_variant_codes.add(_vc_raw)

        variant_code_map: dict[str, str] = {}
        if all_variant_codes:
            for pv in ProductVariant.objects.filter(
                sku_variant_code__in=all_variant_codes,
                company=company,
            ).only("id", "sku_variant_code"):
                variant_code_map[pv.sku_variant_code] = str(pv.id)

        # Pre-load ColorAbbreviation for this company
        color_abbreviations = set(
            ColorAbbreviation.objects.filter(company=company).values_list("color_name", flat=True)
        )

        valid: list[dict] = []
        errors: list[dict] = []
        missing_colors_list: list[dict] = []
        missing_product_names_list: list[dict] = []
        dim_mismatches_list: list[dict] = []

        def get_cell(row: tuple, col: str) -> str | None:
            idx = col_index.get(col)
            if idx is None or idx >= len(row):
                return None
            val = row[idx]
            return str(val).strip() if val is not None else None

        # First pass: collect per-row data for cross-row checks
        row_data_list: list[dict] = []

        for row_num, row in enumerate(rows[1:], start=2):
            product_name = get_cell(row, "product_name") or ""
            variant_code = get_cell(row, "variant_code") or ""
            category_code = get_cell(row, "category_code") or ""
            unit_price_raw = get_cell(row, "unit_price")
            supplier_link_raw = get_cell(row, "supplier_link") or ""
            dim1_key_val = get_cell(row, "dim1_key") or ""
            dim1_value_val = get_cell(row, "dim1_value") or ""
            dim2_key_val = get_cell(row, "dim2_key") or ""
            dim2_value_val = get_cell(row, "dim2_value") or ""

            # Blank-row detection
            if not any(
                [
                    product_name,
                    dim1_value_val,
                    dim2_value_val,
                    unit_price_raw,
                    variant_code,
                    supplier_link_raw,
                    get_cell(row, "image_url"),
                ]
            ):
                continue

            row_errors: list[str] = []

            # Error 1: dim1_key filled but dim1_value empty
            if dim1_key_val and not dim1_value_val:
                row_errors.append("dim1_key is filled but dim1_value is empty")

            # Error 2: dim2_key filled but dim2_value empty
            if dim2_key_val and not dim2_value_val:
                row_errors.append("dim2_key is filled but dim2_value is empty")

            # Error 3: product_name AND supplier_link AND variant_code all blank
            if not product_name and not supplier_link_raw and not variant_code:
                row_errors.append("product_name or supplier_link is required")

            # Derive variant_name for error checking
            if dim1_value_val and dim2_value_val:
                derived_variant_name = f"{dim1_value_val}-{dim2_value_val}"
            elif dim1_value_val:
                derived_variant_name = dim1_value_val
            elif dim2_value_val:
                derived_variant_name = dim2_value_val
            else:
                derived_variant_name = ""

            unit_price: Decimal | None = None
            # Error 4: unit_price required
            if not unit_price_raw:
                row_errors.append("unit_price is required")
            else:
                try:
                    unit_price = Decimal(unit_price_raw)
                    if unit_price <= 0:
                        row_errors.append("unit_price must be greater than zero")
                except InvalidOperation:
                    row_errors.append(f"unit_price '{unit_price_raw}' is not a valid number")

            if row_errors:
                errors.append({"row": row_num, "message": "; ".join(row_errors)})
                continue

            discounted_price: Decimal | None = None
            discounted_price_raw_val = get_cell(row, "discounted_price")
            if discounted_price_raw_val:
                try:
                    discounted_price = Decimal(discounted_price_raw_val)
                    if discounted_price <= 0:
                        errors.append(
                            {
                                "row": row_num,
                                "message": "discounted_price must be greater than zero",
                            }
                        )
                        continue
                except InvalidOperation:
                    errors.append(
                        {
                            "row": row_num,
                            "message": f"discounted_price '{discounted_price_raw_val}' is not a valid number",
                        }
                    )
                    continue

            assert unit_price is not None

            # discounted_price == unit_price → silently set to None
            if discounted_price is not None and discounted_price >= unit_price:
                if discounted_price == unit_price:
                    discounted_price = None
                else:
                    errors.append(
                        {
                            "row": row_num,
                            "message": "discounted_price must be less than unit_price",
                        }
                    )
                    continue

            order_qty_raw_val = get_cell(row, "order_qty") or get_cell(row, "qty_suggested")
            qty_suggested: int | None = None
            if order_qty_raw_val:
                try:
                    qty_suggested = int(Decimal(str(order_qty_raw_val)))
                    if qty_suggested < 0:
                        errors.append({"row": row_num, "message": "order_qty must be 0 or greater"})
                        continue
                except (ValueError, InvalidOperation):
                    errors.append(
                        {
                            "row": row_num,
                            "message": f"order_qty '{order_qty_raw_val}' must be a whole number",
                        }
                    )
                    continue

            # variant_code not found in DB → no longer an error, pass through
            variant_id: str | None = None
            if variant_code:
                variant_id = variant_code_map.get(variant_code)

            cat_info = category_map.get(category_code) if category_code else None

            row_data = {
                "row": row_num,
                "product_name": product_name or None,
                "variant_name": derived_variant_name,
                "variant_code": variant_code or None,
                "variant_id": variant_id,
                "category_code": category_code or None,
                "category_id": cat_info["id"] if cat_info else None,
                "category_name": cat_info["name"] if cat_info else None,
                "unit_price": str(unit_price),
                "discounted_price": str(discounted_price) if discounted_price is not None else None,
                "qty_suggested": qty_suggested,
                "supplier_link": supplier_link_raw or None,
                "image_url": get_cell(row, "image_url"),
                "notes": get_cell(row, "notes"),
                "dim1_key": dim1_key_val or None,
                "dim1_value": dim1_value_val or None,
                "dim2_key": dim2_key_val or None,
                "dim2_value": dim2_value_val or None,
            }

            valid.append(row_data)

            # Detect missing_colors
            color_like_keys = {"color", "warna", "colour"}
            if dim1_key_val and dim1_key_val.lower() in color_like_keys and dim1_value_val:
                if dim1_value_val not in color_abbreviations:
                    missing_colors_list.append({"color_name": dim1_value_val})
            if dim2_key_val and dim2_key_val.lower() in color_like_keys and dim2_value_val:
                if dim2_value_val not in color_abbreviations:
                    missing_colors_list.append({"color_name": dim2_value_val})

            # Detect missing_product_names
            if not product_name and not supplier_link_raw and not variant_code:
                missing_product_names_list.append(
                    {
                        "row": row_num,
                        "supplier_link": None,
                        "dim1_key": dim1_key_val or None,
                        "dim1_value": dim1_value_val or None,
                        "dim2_key": dim2_key_val or None,
                        "dim2_value": dim2_value_val or None,
                        "unit_price": str(unit_price),
                    }
                )

            # Detect dim_mismatches
            if variant_code and dim1_value_val:
                dim_mismatches_list.append(
                    {
                        "row": row_num,
                        "variant_code": variant_code,
                        "dim1_key": dim1_key_val or None,
                        "dim1_value": dim1_value_val or None,
                        "dim2_key": dim2_key_val or None,
                        "dim2_value": dim2_value_val or None,
                    }
                )

            row_data_list.append(row_data)

        # Cross-row checks (pass 2)
        product_groups: dict[str, list[dict]] = {}
        for rd in row_data_list:
            pn = rd["product_name"]
            if pn:
                pn_lower = pn.lower()
                if pn_lower not in product_groups:
                    product_groups[pn_lower] = []
                product_groups[pn_lower].append(rd)

        for pn_lower, group_rows in product_groups.items():
            if len(group_rows) < 2:
                continue
            # Check dim1_key consistency
            dim1_keys = {r["dim1_key"] for r in group_rows if r["dim1_key"]}
            if len(dim1_keys) > 1:
                errors.append(
                    {
                        "row": group_rows[1]["row"],
                        "message": f"Inconsistent dim1_key for product '{group_rows[0]['product_name']}': {', '.join(sorted(dim1_keys))}",
                    }
                )
            # Check dim2_key consistency
            dim2_keys = {r["dim2_key"] for r in group_rows if r["dim2_key"]}
            if len(dim2_keys) > 1:
                errors.append(
                    {
                        "row": group_rows[1]["row"],
                        "message": f"Inconsistent dim2_key for product '{group_rows[0]['product_name']}': {', '.join(sorted(dim2_keys))}",
                    }
                )
            # Check category_code consistency
            cat_codes = {r["category_code"] for r in group_rows if r["category_code"]}
            if len(cat_codes) > 1:
                errors.append(
                    {
                        "row": group_rows[1]["row"],
                        "message": f"Inconsistent category_code for product '{group_rows[0]['product_name']}': {', '.join(sorted(cat_codes))}",
                    }
                )

        # Deduplicate missing_colors
        seen_colors: set[str] = set()
        unique_missing_colors: list[dict] = []
        for mc in missing_colors_list:
            cn = mc["color_name"]
            if cn not in seen_colors:
                seen_colors.add(cn)
                unique_missing_colors.append(mc)

        return {
            "valid": valid,
            "errors": errors,
            "missing_colors": unique_missing_colors,
            "missing_product_names": missing_product_names_list,
            "dim_mismatches": dim_mismatches_list,
        }
