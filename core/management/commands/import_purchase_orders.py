"""
Management command: import_purchase_orders

Usage:
    uv run python manage.py import_purchase_orders --file /path/to/workbook.xlsx
    uv run python manage.py import_purchase_orders --file /path/to/workbook.xlsx --company Mirako
"""

from argparse import ArgumentParser
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.purchasing.services.po_import_service import PurchaseOrderImportService
from core.management.commands.import_purchase_orders_parser import parse_po_sheet
from core.models import Company

SHEETS_TO_IMPORT = [
    "PO Recap 2 April 2025",
    "PO Recap 2 Juni 2025",
    "PO Recap 19 Juni 2025",
    "Recap 11 Sept 2025",
    "Recap 13 Oct 2025",
    "Recap 5 Dec 2025",
    "Recap 21 Dec 2025",
    "Recap 9 Jan 2026",
    "Recap 24 Feb 2026",
    "Split Recap 24 Feb 2026",
    "Recap 29 Mar 2026",
    "Split Recap 29 Mar 2026",
    "Recap 12 June 2026",
    "Recap 29 June 2026",
]


class Command(BaseCommand):
    help = "Import historical purchase orders from an Excel workbook"

    def add_arguments(self, parser: ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Path to the Excel workbook (.xlsx)")
        parser.add_argument(
            "--company",
            default="Mirako",
            help="Company name to import into (default: Mirako)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        file_path: str = options["file"]
        company_name: str = options["company"]

        # 1. Resolve company
        try:
            company = Company.objects.get(name=company_name)
        except Company.DoesNotExist:
            raise CommandError(
                f"Company '{company_name}' not found. "
                "Create it first with: manage.py create_user --company <name>"
            )

        # 2. Load workbook
        self.stdout.write(f"Loading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open workbook: {exc}") from exc

        # 3. Resolve warehouse (first warehouse for company, ordered by cdate)
        from apps.inventory.models import Warehouse

        warehouse = Warehouse.objects.filter(company=company).order_by("cdate").first()
        if warehouse is None:
            raise CommandError(
                f"No warehouses found for company '{company_name}'. Run import_master_data first."
            )
        self.stdout.write(f"  Warehouse: {warehouse.name}")

        # 4. Resolve supplier (first supplier for company)
        from apps.purchasing.models import Supplier

        supplier = Supplier.objects.filter(company=company).first()
        if supplier is None:
            raise CommandError(
                f"No suppliers found for company '{company_name}'. Run import_master_data first."
            )
        self.stdout.write(f"  Supplier:  {supplier.name}")

        # 5. Build variant_map: {sku_variant_code: str(variant_id)}
        from apps.catalog.models import ProductVariant

        variant_map: dict[str, str] = {
            row["sku_variant_code"]: str(row["id"])
            for row in ProductVariant.objects.filter(company=company).values(
                "sku_variant_code", "id"
            )
        }
        self.stdout.write(f"  Variants in DB: {len(variant_map)}")

        # 6. Parse each sheet from SHEETS_TO_IMPORT
        parsed_sheets = []
        sheets_skipped = 0

        for sheet_name in SHEETS_TO_IMPORT:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(f"  [SKIP] Sheet not in workbook: {sheet_name!r}")
                sheets_skipped += 1
                continue

            rows = list(wb[sheet_name].iter_rows(values_only=True))
            parsed = parse_po_sheet(sheet_name, rows)
            if parsed is None:
                self.stdout.write(f"  [SKIP] Could not parse sheet: {sheet_name!r}")
                sheets_skipped += 1
            else:
                self.stdout.write(
                    f"  [OK]   {sheet_name!r} → {parsed.po_date} "
                    f"rate={parsed.exchange_rate} rows={len(parsed.rows)}"
                )
                parsed_sheets.append(parsed)

        self.stdout.write(f"\nParsed {len(parsed_sheets)} sheets, skipped {sheets_skipped}")

        if not parsed_sheets:
            raise CommandError("No sheets could be parsed — nothing to import.")

        # 7. Call service
        service = PurchaseOrderImportService()
        result = service.import_purchase_orders(
            company=company,
            warehouse=warehouse,
            supplier=supplier,
            parsed_sheets=parsed_sheets,
            variant_map=variant_map,
        )

        # 8. Print summary
        self.stdout.write(self.style.SUCCESS("\nImport complete:"))
        self.stdout.write(f"  Sheets imported:  {result.sheets_imported}")
        self.stdout.write(f"  Sheets skipped:   {result.sheets_skipped}")
        self.stdout.write(f"  POs created:      {result.pos_created}")
        self.stdout.write(f"  Details created:  {result.details_created}")
        self.stdout.write(f"  COGS created:     {result.cogs_created}")
        self.stdout.write(f"  Movements created:{result.movements_created}")
