"""
Management command: import_master_data

Usage:
    uv run python manage.py import_master_data --file /path/to/workbook.xlsx
    uv run python manage.py import_master_data --file /path/to/workbook.xlsx --company Mirako
"""

from argparse import ArgumentParser
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.catalog.services.catalog_import_service import CatalogImportService
from core.management.commands.import_master_data_parser import (
    parse_master_sku_sheet,
    parse_variant_sheet,
)
from core.models import Company


class Command(BaseCommand):
    help = "Import master catalog data (products, variants, stock) from an Excel workbook"

    def add_arguments(self, parser: ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Path to the Excel workbook (.xlsx)")
        parser.add_argument(
            "--company",
            default="Mirako",
            help="Company name to import into (default: Mirako)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        file_path: str = options["file"]
        company_name: str = options["company"]

        try:
            company = Company.objects.get(name=company_name)
        except Company.DoesNotExist:
            raise CommandError(
                f"Company '{company_name}' not found. "
                "Create it first with: manage.py create_user --company <name>"
            )

        self.stdout.write(f"Loading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open workbook: {exc}") from exc

        try:
            master_sku_raw = list(wb["Master SKU"].iter_rows(values_only=True))
            variant_raw = list(wb["Master SKU Variant"].iter_rows(values_only=True))
        except KeyError as exc:
            raise CommandError(f"Sheet not found in workbook: {exc}") from exc

        self.stdout.write(
            f"  Master SKU rows (incl. header): {len(master_sku_raw)}\n"
            f"  Master SKU Variant rows (incl. header): {len(variant_raw)}"
        )

        try:
            master_sku_rows = parse_master_sku_sheet(master_sku_raw)
            variant_rows = parse_variant_sheet(variant_raw)
        except ValueError as exc:
            raise CommandError(f"Parse error: {exc}") from exc

        self.stdout.write(
            f"  Parsed {len(master_sku_rows)} SKU rows, {len(variant_rows)} variant rows"
        )

        service = CatalogImportService()
        result = service.import_master_data(
            company=company,
            master_sku_rows=master_sku_rows,
            variant_rows=variant_rows,
        )

        self.stdout.write(self.style.SUCCESS("\nImport complete:"))
        self.stdout.write(f"  Suppliers created:        {result.suppliers_created}")
        self.stdout.write(f"  Categories created:       {result.categories_created}")
        self.stdout.write(f"  Products created:         {result.products_created}")
        self.stdout.write(f"  Variants created:         {result.variants_created}")
        self.stdout.write(f"  Warehouse stock rows:     {result.warehouse_stock_created}")
        self.stdout.write(f"  ProductSupplier links:    {result.product_suppliers_created}")
        self.stdout.write(f"  Warehouse:                {result.warehouse_name}")
