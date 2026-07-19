"""
Management command: import_sales

Usage:
    uv run python manage.py import_sales --file /path/to/workbook.xlsx
    uv run python manage.py import_sales --file /path/to/workbook.xlsx --company Mirako
"""

from argparse import ArgumentParser
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.models import Company
from core.parsers.import_sales_parser import parse_sales_sheet

SHEETS_TO_IMPORT = [
    "Sales May 2025",
    "Sales June 2025",
    "Sales July 2025",
    "Sales August 2025",
    "Sales September 2025",
    "Sales October 2025",
    "Sales November 2025",
    "Sales December 2025",
    "Sales January 2026",
    "Sales February 2026",
    "Sales March 2026",
    "Sales April 2026",
    "Sales May 2026",
]


class Command(BaseCommand):
    help = "Import historical sales orders from an Excel workbook"

    def add_arguments(self, parser: ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Path to the Excel workbook (.xlsx)")
        parser.add_argument(
            "--company",
            default="Mirako",
            help="Company name to import into (default: Mirako)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        file_path: str = options["file"]
        company_name: str = options["company"]

        # 1. Resolve company
        try:
            company = Company.objects.get(name=company_name)
        except Company.DoesNotExist:
            raise CommandError(
                f"Company '{company_name}' not found. "
                "Create it first with: manage.py create_user --company <name>"
            )

        # 2. Load workbook
        self.stdout.write(f"Loading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open workbook: {exc}") from exc

        # 3. Resolve warehouse (first warehouse for company, ordered by cdate)
        from apps.inventory.models import Warehouse

        warehouse = Warehouse.objects.filter(company=company).order_by("cdate").first()
        if warehouse is None:
            raise CommandError(
                f"No warehouses found for company '{company_name}'. Run import_master_data first."
            )
        self.stdout.write(f"  Warehouse: {warehouse.name}")

        # 4. Build variant_map: {sku_variant_code: str(variant_id)}
        from apps.catalog.models import ProductVariant

        variant_map: dict[str, str] = {
            row["sku_variant_code"]: str(row["id"])
            for row in ProductVariant.objects.filter(company=company).values(
                "sku_variant_code", "id"
            )
        }
        self.stdout.write(f"  Variants in DB: {len(variant_map)}")

        # 5. Parse each sheet from SHEETS_TO_IMPORT
        all_parsed_orders = []
        sheets_imported = 0
        sheets_skipped = 0

        for sheet_name in SHEETS_TO_IMPORT:
            if sheet_name not in wb.sheetnames:
                self.stdout.write(f"  [SKIP] Sheet not in workbook: {sheet_name!r}")
                sheets_skipped += 1
                continue

            rows = list(wb[sheet_name].iter_rows(values_only=True))
            parsed = parse_sales_sheet(sheet_name, rows)
            if not parsed:
                self.stdout.write(f"  [SKIP] Could not parse sheet: {sheet_name!r}")
                sheets_skipped += 1
            else:
                self.stdout.write(f"  [OK]   {sheet_name!r} → {len(parsed)} orders")
                all_parsed_orders.extend(parsed)
                sheets_imported += 1

        self.stdout.write(
            f"\nParsed {sheets_imported} sheets, skipped {sheets_skipped}, "
            f"total orders: {len(all_parsed_orders)}"
        )

        if not all_parsed_orders:
            raise CommandError("No orders could be parsed — nothing to import.")

        # 6. Call service
        from apps.sales.services.sales_import_service import SalesImportService

        service = SalesImportService()
        result = service.import_sales(
            company=company,
            warehouse=warehouse,
            variant_map=variant_map,
            parsed_orders=all_parsed_orders,
        )

        # 7. Print summary
        self.stdout.write(self.style.SUCCESS("\nImport complete:"))
        self.stdout.write(f"  Sheets imported:          {sheets_imported}")
        self.stdout.write(f"  Sheets skipped:           {sheets_skipped}")
        self.stdout.write(f"  Orders created:           {result.orders_created}")
        self.stdout.write(f"  Items created:            {result.items_created}")
        self.stdout.write(f"  Orders skipped (dupl):    {result.orders_skipped_duplicate}")
        self.stdout.write(f"  Items skipped (no SKU):   {result.items_skipped_no_sku}")
        self.stdout.write(f"  FIFO shortage warnings:   {result.fifo_shortage_warnings}")
