"""
Management command: import_cash_transactions

Usage:
    uv run python manage.py import_cash_transactions --file /path/to/workbook.xlsx
    uv run python manage.py import_cash_transactions --file /path/to/workbook.xlsx --company Mirako
"""

from argparse import ArgumentParser
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from core.management.commands.import_cash_transactions_parser import (
    parse_cash_transactions_sheet,
)
from core.models import Company

SHEET_NAME = "Finance"


class Command(BaseCommand):
    help = "Import cash transactions from an Excel Finance sheet"

    def add_arguments(self, parser: ArgumentParser) -> None:
        parser.add_argument("--file", required=True, help="Path to the Excel workbook (.xlsx)")
        parser.add_argument(
            "--company",
            default="Mirako",
            help="Company name (default: Mirako)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        file_path: str = options["file"]
        company_name: str = options["company"]

        # 1. Resolve company
        try:
            company = Company.objects.get(name=company_name)
        except Company.DoesNotExist:
            raise CommandError(
                f"Company '{company_name}' not found. "
                "Create it first with: manage.py create_user --company <name>"
            )

        # 2. Load workbook
        self.stdout.write(f"Loading workbook: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as exc:
            raise CommandError(f"Cannot open workbook: {exc}") from exc

        # 3. Check sheet exists
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(
                f"Sheet '{SHEET_NAME}' not found in workbook. "
                f"Available sheets: {', '.join(wb.sheetnames)}"
            )

        # 4. Parse sheet
        rows = list(wb[SHEET_NAME].iter_rows(values_only=True))
        self.stdout.write(f"  Sheet '{SHEET_NAME}' has {len(rows)} rows (including header)")

        parsed_rows = parse_cash_transactions_sheet(rows)

        # Compute skip counts (rows that were skipped by the parser)
        data_rows = len(rows) - 1  # subtract header
        parsed_count = len(parsed_rows)
        skipped_total = data_rows - parsed_count

        self.stdout.write(f"  Parsed {parsed_count} valid rows, skipped {skipped_total} rows")

        # 5. Call service
        from apps.finance.services.cash_transaction_import_service import (
            CashTransactionImportService,
        )

        result = CashTransactionImportService().import_cash_transactions(
            company=company,
            parsed_rows=parsed_rows,
        )

        # 6. Print summary
        self.stdout.write(self.style.SUCCESS("\nImport complete:"))
        self.stdout.write(f"  Inserted: {result.inserted}")
        self.stdout.write(f"  Skipped:  {skipped_total}")
